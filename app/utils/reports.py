"""Report generation utilities"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime
from typing import List, Dict
import pytz

IST = pytz.timezone('Asia/Kolkata')

def generate_excel_report(data: List[Dict], headers: List[str], title: str) -> BytesIO:
    """Generate Excel file from data"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated on: {datetime.now(IST).strftime('%d-%m-%Y %I:%M %p IST')}")
    timestamp_cell.alignment = Alignment(horizontal="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_num, row_data in enumerate(data, 5):
        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            ws.cell(row=row_num, column=col_num, value=value)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_report(data: List[Dict], headers: List[str], title: str) -> BytesIO:
    """Generate PDF file from data"""
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=1
    )
    elements.append(Paragraph(title, title_style))
    
    timestamp_text = f"Generated on: {datetime.now(IST).strftime('%d-%m-%Y %I:%M %p IST')}"
    timestamp_style = ParagraphStyle('Timestamp', parent=styles['Normal'], fontSize=9, alignment=1)
    elements.append(Paragraph(timestamp_text, timestamp_style))
    elements.append(Spacer(1, 20))
    
    table_data = [headers]
    for row in data:
        table_data.append([str(row.get(header, "")) for header in headers])
    
    col_widths = [A4[0] / len(headers) - 10] * len(headers)
    table = Table(table_data, colWidths=col_widths)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output
