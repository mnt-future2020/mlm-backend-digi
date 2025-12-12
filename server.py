from fastapi import FastAPI, HTTPException, Depends, status, Body, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, EmailStr, Field, field_validator
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta, timezone
from jose import JWTError, jwt
from passlib.context import CryptContext
from pymongo import MongoClient, ASCENDING, DESCENDING
from bson import ObjectId
import os
from dotenv import load_dotenv
import random
import string
import re
import pytz
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Auto-placement functions (moved from service to avoid import issues)

def find_deepest_left_position(sponsor_id: str):
    """Find the deepest LEFT-most available position in sponsor's LEFT leg"""
    left_child = teams_collection.find_one({
        "sponsorId": sponsor_id,
        "placement": "LEFT"
    })
    
    if not left_child:
        return None
    
    current_user_id = left_child["userId"]
    max_depth = 100
    for _ in range(max_depth):
        left_child = teams_collection.find_one({
            "sponsorId": current_user_id,
            "placement": "LEFT"
        })
        
        if not left_child:
            return current_user_id
        
        current_user_id = left_child["userId"]
    
    return current_user_id

def find_deepest_right_position(sponsor_id: str):
    """Find the deepest RIGHT-most available position in sponsor's RIGHT leg"""
    right_child = teams_collection.find_one({
        "sponsorId": sponsor_id,
        "placement": "RIGHT"
    })
    
    if not right_child:
        return None
    
    current_user_id = right_child["userId"]
    max_depth = 100
    for _ in range(max_depth):
        right_child = teams_collection.find_one({
            "sponsorId": current_user_id,
            "placement": "RIGHT"
        })
        
        if not right_child:
            return current_user_id
        
        current_user_id = right_child["userId"]
    
    return current_user_id

def get_auto_placement_position(sponsor_id: str, preferred_placement: str):
    """Get the actual placement position for a new user"""
    if preferred_placement == "LEFT":
        actual_sponsor = find_deepest_left_position(sponsor_id)
        if actual_sponsor is None:
            return sponsor_id, "LEFT"
        else:
            return actual_sponsor, "LEFT"
    elif preferred_placement == "RIGHT":
        actual_sponsor = find_deepest_right_position(sponsor_id)
        if actual_sponsor is None:
            return sponsor_id, "RIGHT"
        else:
            return actual_sponsor, "RIGHT"
    else:
        return sponsor_id, "LEFT"

def get_placement_info_for_display(sponsor_id: str, preferred_placement: str):
    """Get human-readable placement information for UI display"""
    original_sponsor = users_collection.find_one({"_id": ObjectId(sponsor_id)})
    if not original_sponsor:
        return None
    
    actual_sponsor_id, placement = get_auto_placement_position(sponsor_id, preferred_placement)
    actual_sponsor = users_collection.find_one({"_id": ObjectId(actual_sponsor_id)})
    if not actual_sponsor:
        return None
    
    is_direct = (sponsor_id == actual_sponsor_id)
    
    return {
        "original_sponsor_id": sponsor_id,
        "original_sponsor_name": original_sponsor.get("name", "Unknown"),
        "original_sponsor_referral_id": original_sponsor.get("referralId", "Unknown"),
        "actual_sponsor_id": actual_sponsor_id,
        "actual_sponsor_name": actual_sponsor.get("name", "Unknown"),
        "actual_sponsor_referral_id": actual_sponsor.get("referralId", "Unknown"),
        "placement": placement,
        "is_direct_placement": is_direct,
        "message": f"Will be placed under {actual_sponsor.get('name')} on {placement} side"
    }

# Indian Standard Time timezone
IST = pytz.timezone('Asia/Kolkata')

# Load environment variables
load_dotenv()

# Initialize FastAPI app
app = FastAPI(title="VSV Unite MLM API", version="1.0.0")

# CORS Configuration
cors_origins = os.getenv("CORS_ORIGINS", "*")
if cors_origins == "*":
    allowed_origins = ["*"]
else:
    allowed_origins = cors_origins.split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB Configuration
MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017/")
MONGO_DB_NAME = os.getenv("MONGO_DB_NAME", "mlm_vsv_unite")
client = MongoClient(MONGO_URL)
db = client[MONGO_DB_NAME]

# Collections
users_collection = db["users"]
plans_collection = db["plans"]
wallets_collection = db["wallets"]
transactions_collection = db["transactions"]
teams_collection = db["teams"]
withdrawals_collection = db["withdrawals"]
settings_collection = db["settings"]
email_configs_collection = db["email_configs"]
topups_collection = db["topups"]
ranks_collection = db["ranks"]

# JWT Configuration
JWT_SECRET_KEY = os.getenv("JWT_SECRET_KEY")
JWT_ALGORITHM = os.getenv("JWT_ALGORITHM", "HS256")
JWT_ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("JWT_ACCESS_TOKEN_EXPIRE_MINUTES", 10080))

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Helper functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def get_system_time_offset():
    """Get system time offset from settings (in minutes)"""
    try:
        settings = settings_collection.find_one({})
        if settings and settings.get("systemTimeOffset"):
            return int(settings.get("systemTimeOffset", 0))
    except:
        pass
    return 0

def get_ist_now():
    """Get current time in IST with optional offset from settings"""
    base_time = datetime.now(IST)
    offset_minutes = get_system_time_offset()
    if offset_minutes != 0:
        return base_time + timedelta(minutes=offset_minutes)
    return base_time

def get_eod_time():
    """Get End of Day time from settings (default 23:59)"""
    try:
        settings = settings_collection.find_one({})
        if settings and settings.get("eodTime"):
            return settings.get("eodTime", "23:59")
    except:
        pass
    return "23:59"

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=JWT_ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)
    return encoded_jwt

def generate_referral_id(prefix="VSV"):
    """Generate unique referral ID"""
    while True:
        random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=7))
        referral_id = f"{prefix}{random_str}"
        if not users_collection.find_one({"referralId": referral_id}):
            return referral_id

def get_user_rank(total_pv: int):
    """Get user rank based on total PV"""
    # Get all ranks sorted by minPV descending
    ranks = list(ranks_collection.find({}).sort("minPV", DESCENDING))
    
    # Find the highest rank user qualifies for
    for rank in ranks:
        if total_pv >= rank.get("minPV", 0):
            return {
                "name": rank.get("name"),
                "icon": rank.get("icon"),
                "color": rank.get("color"),
                "minPV": rank.get("minPV")
            }
    
    # If no rank found, return lowest rank
    lowest_rank = ranks_collection.find_one({}, sort=[("minPV", ASCENDING)])
    if lowest_rank:
        return {
            "name": lowest_rank.get("name"),
            "icon": lowest_rank.get("icon"),
            "color": lowest_rank.get("color"),
            "minPV": lowest_rank.get("minPV")
        }
    
    # Default rank if no ranks defined
    return {
        "name": "Member",
        "icon": "👤",
        "color": "#6B7280",
        "minPV": 0
    }

def serialize_doc(doc):
    """Convert MongoDB document to JSON serializable format"""
    if doc is None:
        return None
    if isinstance(doc, list):
        return [serialize_doc(item) for item in doc]
    if isinstance(doc, dict):
        result = {}
        for key, value in doc.items():
            if key == "_id":
                result["id"] = str(value)
            elif isinstance(value, ObjectId):
                result[key] = str(value)
            elif isinstance(value, datetime):
                result[key] = value.isoformat()
            elif isinstance(value, dict):
                result[key] = serialize_doc(value)
            elif isinstance(value, list):
                result[key] = [serialize_doc(item) for item in value]
            else:
                result[key] = value
        return result
    return doc

# ============ REPORT GENERATION HELPERS ============

def generate_excel_report(data: List[Dict], headers: List[str], title: str) -> BytesIO:
    """Generate Excel file from data"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add timestamp
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated on: {datetime.now(IST).strftime('%d-%m-%Y %I:%M %p IST')}")
    timestamp_cell.alignment = Alignment(horizontal="center")
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data
    for row_num, row_data in enumerate(data, 5):
        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_report(data: List[Dict], headers: List[str], title: str) -> BytesIO:
    """Generate PDF file from data"""
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph(title, title_style))
    
    # Timestamp
    timestamp_text = f"Generated on: {datetime.now(IST).strftime('%d-%m-%Y %I:%M %p IST')}"
    timestamp_style = ParagraphStyle('Timestamp', parent=styles['Normal'], fontSize=9, alignment=1)
    elements.append(Paragraph(timestamp_text, timestamp_style))
    elements.append(Spacer(1, 20))
    
    # Prepare table data
    table_data = [headers]
    for row in data:
        table_data.append([str(row.get(header, "")) for header in headers])
    
    # Create table
    col_widths = [A4[0] / len(headers) - 10] * len(headers)
    table = Table(table_data, colWidths=col_widths)
    
    # Table styling
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    return output

def parse_date_range(start_date: Optional[str], end_date: Optional[str]):
    """Parse and validate date range parameters"""
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").replace(hour=0, minute=0, second=0, microsecond=0)
        except:
            raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD")
    else:
        start = None
    
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, microsecond=999999)
        except:
            raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD")
    else:
        end = None
    
    return start, end

# Get current user from token
async def get_current_user(authorization: Optional[str] = None):
    """Extract user from JWT token in Authorization header"""
    from fastapi import Header
    
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    token = authorization.replace("Bearer ", "")
    
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
        user_id: str = payload.get("userId")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = users_collection.find_one({"_id": ObjectId(user_id)})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    
    return serialize_doc(user)

async def get_current_active_user(authorization: Optional[str] = Header(None)):
    """Get current active user"""
    user = await get_current_user(authorization)
    if not user.get("isActive"):
        raise HTTPException(status_code=400, detail="Inactive user")
    return user

async def get_current_admin(authorization: Optional[str] = Header(None)):
    """Get current admin user"""
    user = await get_current_user(authorization)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    return user

# Pydantic Models
class UserRegister(BaseModel):
    name: str
    username: str
    email: Optional[str] = None  # Changed from EmailStr to str to allow empty
    password: str
    mobile: str
    referralId: Optional[str] = None
    placement: Optional[str] = None  # LEFT or RIGHT
    planId: Optional[str] = None  # Optional plan assignment during registration
    
    @field_validator('placement')
    @classmethod
    def validate_placement(cls, v):
        if v and v not in ['LEFT', 'RIGHT']:
            raise ValueError('Placement must be LEFT or RIGHT')
        return v
    
    @field_validator('email', mode='before')
    @classmethod
    def validate_email(cls, v):
        # Allow empty string or None - convert to None
        if v is None or v == '' or (isinstance(v, str) and v.strip() == ''):
            return None
        return v

class UserLogin(BaseModel):
    email: Optional[str] = None
    username: Optional[str] = None
    password: str

class ReferralLookup(BaseModel):
    referralId: str

class PasswordChange(BaseModel):
    oldPassword: str
    newPassword: str

class PlanActivation(BaseModel):
    planId: str
    paymentProof: Optional[str] = None

class WithdrawalRequest(BaseModel):
    amount: float
    bankDetails: Dict[str, Any]

class SettingsUpdate(BaseModel):
    companyName: Optional[str] = None
    companyEmail: Optional[str] = None
    companyPhone: Optional[str] = None
    companyAddress: Optional[str] = None
    companyDescription: Optional[str] = None
    metaTitle: Optional[str] = None
    metaDescription: Optional[str] = None
    metaKeywords: Optional[str] = None
    ogImage: Optional[str] = None
    heroBadge: Optional[str] = None
    heroSlides: Optional[List[Dict[str, Any]]] = None

# Initialize default plans
def initialize_plans():
    """Initialize membership plans if they don't exist"""
    existing_plans = plans_collection.count_documents({})
    if existing_plans == 0:
        plans = [
            {
                "name": "Basic",
                "amount": 111,
                "pv": 1,
                "referralIncome": 25,
                "dailyCapping": 250,
                "matchingIncome": 25,
                "description": "Start small, earn steady",
                "features": [
                    "Income Start: ₹25",
                    "Daily Capping: ₹250",
                    "Binary: Left 10 - Right 10 = ₹250",
                    "Basic Support"
                ],
                "isActive": True,
                "createdAt": get_ist_now()
            },
            {
                "name": "Standard",
                "amount": 599,
                "pv": 2,
                "referralIncome": 50,
                "dailyCapping": 500,
                "matchingIncome": 50,
                "description": "Popular choice for growth",
                "features": [
                    "Income Start: ₹50",
                    "Daily Capping: ₹500",
                    "Binary: Left 10 - Right 10 = ₹500",
                    "Standard Support"
                ],
                "isActive": True,
                "popular": True,
                "createdAt": get_ist_now()
            },
            {
                "name": "Advanced",
                "amount": 1199,
                "pv": 4,
                "referralIncome": 100,
                "dailyCapping": 1000,
                "matchingIncome": 100,
                "description": "Accelerate your earnings",
                "features": [
                    "Income Start: ₹100",
                    "Daily Capping: ₹1000",
                    "Binary: Left 10 - Right 10 = ₹1000",
                    "Priority Support"
                ],
                "isActive": True,
                "createdAt": get_ist_now()
            },
            {
                "name": "Premium",
                "amount": 1799,
                "pv": 6,
                "referralIncome": 150,
                "dailyCapping": 1500,
                "matchingIncome": 150,
                "description": "Maximum earning potential",
                "features": [
                    "Income Start: ₹150",
                    "Daily Capping: ₹1500",
                    "Binary: Left 10 - Right 10 = ₹1500",
                    "VIP Support"
                ],
                "isActive": True,
                "createdAt": get_ist_now()
            }
        ]
        plans_collection.insert_many(plans)
        print("✅ Default plans initialized")

# Initialize default ranks
def initialize_ranks():
    """Initialize default ranks if they don't exist"""
    existing_ranks = ranks_collection.count_documents({})
    if existing_ranks == 0:
        default_ranks = [
            {
                "name": "Bronze",
                "minPV": 0,
                "color": "#CD7F32",
                "icon": "🥉",
                "benefits": ["Access to basic features", "Basic support"],
                "order": 1
            },
            {
                "name": "Silver",
                "minPV": 100,
                "color": "#C0C0C0",
                "icon": "🥈",
                "benefits": ["Priority support", "Monthly bonus", "Team building tools"],
                "order": 2
            },
            {
                "name": "Gold",
                "minPV": 500,
                "color": "#FFD700",
                "icon": "🥇",
                "benefits": ["Premium support", "Leadership bonus", "Advanced training"],
                "order": 3
            },
            {
                "name": "Platinum",
                "minPV": 1000,
                "color": "#E5E4E2",
                "icon": "💎",
                "benefits": ["VIP support", "Car fund", "International trips", "Recognition awards"],
                "order": 4
            },
            {
                "name": "Diamond",
                "minPV": 5000,
                "color": "#B9F2FF",
                "icon": "💍",
                "benefits": ["Elite support", "House fund", "Luxury trips", "Leadership summit"],
                "order": 5
            }
        ]
        ranks_collection.insert_many(default_ranks)
        print("✅ Default ranks initialized")

# Initialize admin user
def initialize_admin():
    """Create admin user if not exists"""
    admin_email = os.getenv("ADMIN_EMAIL", "admin@vsvunite.com")
    admin_user = users_collection.find_one({"email": admin_email})
    
    if not admin_user:
        admin_password = os.getenv("ADMIN_PASSWORD", "Admin@123")
        admin_referral_id = os.getenv("ADMIN_REFERRAL_ID", "VSV00001")
        
        admin_data = {
            "name": os.getenv("ADMIN_NAME", "VSV Admin"),
            "username": os.getenv("ADMIN_USERNAME", "vsvadmin"),
            "email": admin_email,
            "password": hash_password(admin_password),
            "mobile": "8807867028",
            "referralId": admin_referral_id,
            "role": "admin",
            "isActive": True,
            "isEmailVerified": True,
            "placement": None,
            "sponsorId": admin_referral_id,  # Self-reference for root user
            "currentPlan": None,
            "totalPV": 0,
            "leftPV": 0,
            "rightPV": 0,
            "createdAt": get_ist_now(),
            "updatedAt": get_ist_now()
        }
        
        result = users_collection.insert_one(admin_data)
        
        # Create admin wallet
        wallets_collection.insert_one({
            "userId": str(result.inserted_id),
            "balance": 0,
            "totalEarnings": 0,
            "totalWithdrawals": 0,
            "createdAt": get_ist_now(),
            "updatedAt": get_ist_now()
        })
        
        print(f"✅ Admin user created - Email: {admin_email}, Password: {admin_password}")

# Startup event
@app.on_event("startup")
async def startup_event():
    """Initialize database on startup"""
    # Create indexes - drop existing email index and recreate with sparse
    try:
        users_collection.drop_index("email_1")
    except:
        pass  # Index may not exist
    
    users_collection.create_index([("email", ASCENDING)], unique=True, sparse=True, name="email_1")
    users_collection.create_index([("username", ASCENDING)], unique=True)
    users_collection.create_index([("referralId", ASCENDING)], unique=True)
    users_collection.create_index([("mobile", ASCENDING)])
    
    wallets_collection.create_index([("userId", ASCENDING)], unique=True)
    transactions_collection.create_index([("userId", ASCENDING)])
    teams_collection.create_index([("userId", ASCENDING)])
    teams_collection.create_index([("sponsorId", ASCENDING)])
    
    # Initialize data
    initialize_plans()
    initialize_ranks()
    initialize_admin()
    
    print("✅ Database initialized successfully")

# ==================== AUTH ROUTES ====================

@app.post("/api/auth/register")
async def register(user: UserRegister):
    """Register new user with MLM structure"""
    try:
        # Check if user already exists
        if user.email and users_collection.find_one({"email": user.email}):
            raise HTTPException(status_code=400, detail="Email already registered")
        
        if users_collection.find_one({"username": user.username}):
            raise HTTPException(status_code=400, detail="Username already taken")
        
        # Validate referral ID if provided
        sponsor = None
        actual_sponsor_id = None
        actual_placement = None
        
        if user.referralId:
            sponsor = users_collection.find_one({"referralId": user.referralId})
            if not sponsor:
                raise HTTPException(status_code=400, detail="Invalid referral ID")
            
            if not user.placement:
                raise HTTPException(status_code=400, detail="Placement is required when using referral ID")
            
            # Get auto-placement position (deepest left-most or right-most)
            actual_sponsor_id, actual_placement = get_auto_placement_position(
                str(sponsor["_id"]), 
                user.placement
            )
        
        # Generate unique referral ID
        referral_id = generate_referral_id()
        
        # Check if plan is provided and valid
        plan = None
        if user.planId:
            plan = plans_collection.find_one({"_id": ObjectId(user.planId)})
            if not plan:
                raise HTTPException(status_code=400, detail="Invalid plan ID")
        
        # Create user - only include email if provided (for sparse index to work)
        user_data = {
            "name": user.name,
            "username": user.username,
            "password": hash_password(user.password),
            "mobile": user.mobile,
            "referralId": referral_id,
            "role": "user",
            "isActive": True,
            "isEmailVerified": False,
            "placement": user.placement,
            "sponsorId": user.referralId,
            "currentPlan": plan["name"] if plan else None,
            "currentPlanId": user.planId if plan else None,
            "activatedAt": None,
            "totalPV": 0,
            "leftPV": 0,
            "rightPV": 0,
            "createdAt": get_ist_now(),
            "updatedAt": get_ist_now()
        }
        
        # Only add email field if it has a value (for sparse unique index)
        if user.email:
            user_data["email"] = user.email
        
        result = users_collection.insert_one(user_data)
        user_id = str(result.inserted_id)
        
        # Create wallet
        wallets_collection.insert_one({
            "userId": user_id,
            "balance": 0,
            "totalEarnings": 0,
            "totalWithdrawals": 0,
            "createdAt": get_ist_now(),
            "updatedAt": get_ist_now()
        })
        
        # Add to team structure if has sponsor
        if sponsor:
            # Use auto-placement: actual_sponsor_id and actual_placement
            teams_collection.insert_one({
                "userId": user_id,
                "sponsorId": actual_sponsor_id,  # This is the actual sponsor after auto-placement
                "placement": actual_placement,    # This is the actual placement side
                "level": 1,
                "createdAt": get_ist_now()
            })
            
            # Distribute PV if plan is assigned (referral income system removed)
            if plan:
                # Distribute PV upward in binary tree
                pv_amount = plan.get("pv", 0)
                if pv_amount > 0:
                    distribute_pv_upward(user_id, pv_amount)
                
                # REFERRAL INCOME REMOVED - No longer giving referral income to sponsor
                # referral_income = plan.get("referralIncome", 0)
                # if referral_income > 0:
                #     # Update sponsor wallet
                #     wallets_collection.update_one(
                #         {"userId": str(sponsor["_id"])},
                #         {"$inc": {"balance": referral_income, "totalEarnings": referral_income}}
                #     )
                #     
                #     # Create transaction for sponsor
                #     transactions_collection.insert_one({
                #         "userId": str(sponsor["_id"]),
                #         "type": "REFERRAL_INCOME",
                #         "amount": referral_income,
                #         "description": f"Referral income from {user.name} plan activation",
                #         "status": "COMPLETED",
                #         "createdAt": get_ist_now()
                #     })
        
        # Create access token
        access_token = create_access_token(data={"sub": user.username, "userId": user_id})
        
        # Get created user
        created_user = users_collection.find_one({"_id": result.inserted_id})
        user_response = serialize_doc(created_user)
        user_response.pop("password", None)
        
        return {
            "success": True,
            "message": "Registration successful",
            "user": user_response,
            "token": access_token
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Registration error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/auth/sign-in/email")
async def login_email(credentials: dict = Body(...)):
    """Login with email and password"""
    try:
        email = credentials.get("email")
        password = credentials.get("password")
        
        if not email or not password:
            raise HTTPException(status_code=400, detail="Email and password required")
        
        # Find user
        user = users_collection.find_one({"email": email})
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        # Verify password
        if not verify_password(password, user["password"]):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        # Check if active
        if not user.get("isActive", False):
            raise HTTPException(status_code=403, detail="Account is inactive")
        
        # Create token
        user_id = str(user["_id"])
        access_token = create_access_token(data={"sub": user["username"], "userId": user_id})
        
        # Prepare response
        user_response = serialize_doc(user)
        user_response.pop("password", None)
        
        return {
            "user": user_response,
            "token": access_token,
            "session": {"token": access_token}
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Login error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/auth/sign-in/username")
async def login_username(credentials: dict = Body(...)):
    """Login with username and password"""
    try:
        username = credentials.get("username")
        password = credentials.get("password")
        
        if not username or not password:
            raise HTTPException(status_code=400, detail="Username and password required")
        
        # Find user
        user = users_collection.find_one({"username": username})
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        # Verify password
        if not verify_password(password, user["password"]):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        # Check if active
        if not user.get("isActive", False):
            raise HTTPException(status_code=403, detail="Account is inactive")
        
        # Create token
        user_id = str(user["_id"])
        access_token = create_access_token(data={"sub": user["username"], "userId": user_id})
        
        # Prepare response
        user_response = serialize_doc(user)
        user_response.pop("password", None)
        
        return {
            "user": user_response,
            "token": access_token,
            "session": {"token": access_token}
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"Login error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/auth/lookup-referral")
async def lookup_referral(data: ReferralLookup):
    """Lookup user by referral ID"""
    try:
        user = users_collection.find_one({"referralId": data.referralId})
        if not user:
            return {"success": False, "message": "Invalid referral ID"}
        
        return {
            "success": True,
            "data": {
                "username": user["username"],
                "email": user.get("email", ""),
                "name": user["name"]
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/auth/preview-placement")
async def preview_placement(data: dict = Body(...)):
    """
    Preview where a new user will be placed in the binary tree
    Used to show auto-placement position before registration
    """
    try:
        referral_id = data.get("referralId")
        placement = data.get("placement")  # "LEFT" or "RIGHT"
        
        if not referral_id or not placement:
            raise HTTPException(status_code=400, detail="referralId and placement are required")
        
        # Find sponsor
        sponsor = users_collection.find_one({"referralId": referral_id})
        if not sponsor:
            raise HTTPException(status_code=404, detail="Invalid referral ID")
        
        # Get placement info
        placement_info = get_placement_info_for_display(str(sponsor["_id"]), placement)
        
        if not placement_info:
            raise HTTPException(status_code=500, detail="Could not determine placement")
        
        return {
            "success": True,
            "data": placement_info
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/auth/get-session")
async def get_session():
    """Get current session - placeholder"""
    # In real implementation, you'd validate JWT token from cookie/header
    return {"user": None}

@app.post("/api/auth/sign-out")
async def logout():
    """Logout user"""
    return {"success": True, "message": "Logged out successfully"}

# ==================== USER ROUTES ====================

@app.get("/api/user/profile")
async def get_profile(current_user: dict = Depends(get_current_active_user)):
    """Get user profile"""
    try:
        user_data = serialize_doc(current_user)
        user_data.pop("password", None)
        
        # Get wallet info
        wallet = wallets_collection.find_one({"userId": current_user["id"]})
        if wallet:
            user_data["wallet"] = serialize_doc(wallet)
        
        # Get team count
        team_count = teams_collection.count_documents({"sponsorId": current_user["id"]})
        user_data["teamSize"] = team_count
        
        # Get left and right team counts
        left_count = teams_collection.count_documents({
            "sponsorId": current_user["id"],
            "placement": "LEFT"
        })
        right_count = teams_collection.count_documents({
            "sponsorId": current_user["id"],
            "placement": "RIGHT"
        })
        user_data["leftTeamSize"] = left_count
        user_data["rightTeamSize"] = right_count
        
        return {"success": True, "data": user_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/user/profile")
async def update_profile(
    data: dict = Body(...),
    current_user: dict = Depends(get_current_active_user)
):
    """Update user profile"""
    try:
        # Fields that can be updated
        allowed_fields = ["name", "mobile", "email"]
        update_data = {k: v for k, v in data.items() if k in allowed_fields}
        update_data["updatedAt"] = get_ist_now()
        
        users_collection.update_one(
            {"_id": ObjectId(current_user["id"])},
            {"$set": update_data}
        )
        
        return {"success": True, "message": "Profile updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/user/change-password")
async def change_password(
    data: dict = Body(...),
    current_user: dict = Depends(get_current_active_user)
):
    """Change user password"""
    try:
        old_password = data.get("oldPassword")
        new_password = data.get("newPassword")
        
        if not old_password or not new_password:
            raise HTTPException(status_code=400, detail="Old and new password required")
        
        # Get user from database
        user = users_collection.find_one({"_id": ObjectId(current_user["id"])})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Verify old password
        if not verify_password(old_password, user["password"]):
            raise HTTPException(status_code=400, detail="Incorrect old password")
        
        # Update password
        users_collection.update_one(
            {"_id": ObjectId(current_user["id"])},
            {"$set": {
                "password": hash_password(new_password),
                "updatedAt": get_ist_now()
            }}
        )
        
        return {"success": True, "message": "Password changed successfully"}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/user/referral/{referral_id}")
async def get_referral_info(referral_id: str):
    """Get referral user information"""
    try:
        user = users_collection.find_one({"referralId": referral_id})
        if not user:
            raise HTTPException(status_code=404, detail="Referral ID not found")
        
        return {
            "success": True,
            "data": {
                "name": user["name"],
                "referralId": user["referralId"],
                "isActive": user.get("isActive", False)
            }
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/user/dashboard")
async def get_user_dashboard(current_user: dict = Depends(get_current_active_user)):
    """Get user dashboard statistics"""
    try:
        user_id = current_user["id"]
        
        # Get wallet
        wallet = wallets_collection.find_one({"userId": user_id})
        wallet_data = serialize_doc(wallet) if wallet else {
            "balance": 0,
            "totalEarnings": 0,
            "totalWithdrawals": 0
        }
        
        # Get team statistics
        total_team = teams_collection.count_documents({"sponsorId": user_id})
        left_team = teams_collection.count_documents({
            "sponsorId": user_id,
            "placement": "LEFT"
        })
        right_team = teams_collection.count_documents({
            "sponsorId": user_id,
            "placement": "RIGHT"
        })
        
        # Get current plan (fetch fresh from database, not from JWT token)
        fresh_user = users_collection.find_one({"_id": ObjectId(user_id)})
        current_plan = None
        
        if fresh_user and fresh_user.get("currentPlan"):
            plan_value = fresh_user.get("currentPlan")
            try:
                # Try as ObjectId first
                plan = plans_collection.find_one({"_id": ObjectId(plan_value)})
                if plan:
                    current_plan = serialize_doc(plan)
            except:
                # Try as plan name
                plan = plans_collection.find_one({"name": plan_value})
                if plan:
                    current_plan = serialize_doc(plan)
        
        # Get recent transactions (exclude PLAN_ACTIVATION)
        transactions = list(transactions_collection.find({
            "userId": user_id,
            "type": {"$ne": "PLAN_ACTIVATION"}
        }).sort("createdAt", DESCENDING).limit(5))
        
        # Get user rank based on total PV
        total_pv = fresh_user.get("totalPV", 0) if fresh_user else 0
        user_rank = get_user_rank(total_pv)
        
        return {
            "success": True,
            "data": {
                "wallet": wallet_data,
                "team": {
                    "total": total_team,
                    "left": left_team,
                    "right": right_team
                },
                "currentPlan": current_plan,
                "rank": user_rank,
                "recentTransactions": serialize_doc(transactions)
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/user/team/tree")
async def get_team_tree(current_user: dict = Depends(get_current_active_user)):
    """Get user's team tree (binary structure)"""
    try:
        user_id = current_user["id"]
        
        def build_tree(parent_id, depth=0, max_depth=5):
            if depth > max_depth:
                return None
            
            user = users_collection.find_one({"_id": ObjectId(parent_id)})
            if not user:
                return None
            
            # Get children from teams collection
            # sponsorId in teams collection is stored as string of ObjectId
            left_child = teams_collection.find_one({
                "sponsorId": str(user["_id"]),
                "placement": "LEFT"
            })
            right_child = teams_collection.find_one({
                "sponsorId": str(user["_id"]),
                "placement": "RIGHT"
            })
            
            # Get plan name if exists
            plan_name = None
            if user.get("currentPlan"):
                try:
                    plan = plans_collection.find_one({"_id": ObjectId(user["currentPlan"])}, {"_id": 0})
                    if plan:
                        plan_name = plan.get("name")
                except Exception:
                    # If currentPlan is already a string (plan name), use it
                    plan_name = user.get("currentPlan") if isinstance(user.get("currentPlan"), str) and len(user.get("currentPlan")) < 50 else None
            
            node = {
                "id": str(user["_id"]),
                "name": user["name"],
                "referralId": user["referralId"],
                "placement": user.get("placement"),
                "currentPlan": plan_name,
                "isActive": user.get("isActive", False),
                "leftPV": user.get("leftPV", 0),
                "rightPV": user.get("rightPV", 0),
                "totalPV": user.get("totalPV", 0),
                "left": None,
                "right": None
            }
            
            if left_child:
                node["left"] = build_tree(left_child["userId"], depth + 1, max_depth)
            
            if right_child:
                node["right"] = build_tree(right_child["userId"], depth + 1, max_depth)
            
            return node
        
        tree = build_tree(user_id)
        
        return {
            "success": True,
            "data": tree
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/user/details/{user_id}")
async def get_user_details(user_id: str, current_user: dict = Depends(get_current_active_user)):
    """Get detailed user information"""
    try:
        # Find user by either MongoDB _id or referralId
        try:
            user = users_collection.find_one({"_id": ObjectId(user_id)})
        except:
            user = users_collection.find_one({"referralId": user_id})
        
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Get plan details
        plan_details = None
        if user.get("currentPlan"):
            try:
                plan = plans_collection.find_one({"_id": ObjectId(user["currentPlan"])})
                if plan:
                    plan_details = {
                        "name": plan.get("name"),
                        "amount": plan.get("amount"),
                        "pv": plan.get("pv"),
                        "dailyCapping": plan.get("dailyCapping")
                    }
            except:
                pass
        
        # Get wallet info
        wallet = wallets_collection.find_one({"userId": str(user["_id"])})
        wallet_data = {
            "balance": wallet.get("balance", 0) if wallet else 0,
            "totalEarnings": wallet.get("totalEarnings", 0) if wallet else 0,
            "totalWithdrawals": wallet.get("totalWithdrawals", 0) if wallet else 0
        }
        
        # Get sponsor info
        sponsor_info = None
        if user.get("sponsorId") and user.get("sponsorId") != user.get("referralId"):
            sponsor = users_collection.find_one({"referralId": user["sponsorId"]})
            if sponsor:
                sponsor_info = {
                    "name": sponsor.get("name"),
                    "referralId": sponsor.get("referralId")
                }
        
        # Get team count
        team_count = teams_collection.count_documents({"sponsorId": str(user["_id"])})
        left_count = teams_collection.count_documents({"sponsorId": str(user["_id"]), "placement": "LEFT"})
        right_count = teams_collection.count_documents({"sponsorId": str(user["_id"]), "placement": "RIGHT"})
        
        # Get user's own placement from teams collection
        user_team_record = teams_collection.find_one({"userId": str(user["_id"])})
        user_placement = user_team_record.get("placement") if user_team_record else None
        
        # Get income breakdown from transactions
        income_breakdown = {
            "REFERRAL_INCOME": 0,
            "MATCHING_INCOME": 0,
            "LEVEL_INCOME": 0
        }
        
        income_types = ["REFERRAL_INCOME", "MATCHING_INCOME", "LEVEL_INCOME"]
        for income_type in income_types:
            result = transactions_collection.aggregate([
                {"$match": {"userId": str(user["_id"]), "type": income_type, "status": "COMPLETED"}},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ])
            total = list(result)
            if total:
                income_breakdown[income_type] = total[0].get("total", 0)
        
        # Build response
        user_details = {
            "id": str(user["_id"]),
            "name": user.get("name"),
            "username": user.get("username"),
            "email": user.get("email"),
            "mobile": user.get("mobile"),
            "referralId": user.get("referralId"),
            "sponsorId": user.get("sponsorId"),
            "placement": user_placement,
            "sponsor": sponsor_info,
            "isActive": user.get("isActive", False),
            "currentPlan": plan_details,
            "wallet": wallet_data,
            "incomeBreakdown": income_breakdown,
            "pv": {
                "leftPV": user.get("leftPV", 0),
                "rightPV": user.get("rightPV", 0),
                "totalPV": user.get("totalPV", 0),
                "planPV": plan_details.get("pv", 0) if plan_details else 0,  # PV from user's plan
                "dailyPVUsed": user.get("dailyPVUsed", 0)
            },
            "team": {
                "total": team_count,
                "left": left_count,
                "right": right_count
            },
            "joinedAt": user.get("createdAt"),
            "lastActive": user.get("updatedAt")
        }
        
        return {
            "success": True,
            "data": user_details
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/user/team/list")
async def get_team_list(current_user: dict = Depends(get_current_active_user)):
    """Get user's team list"""
    try:
        user_id = current_user["id"]
        
        # Get all team members
        team_members = list(teams_collection.find({"sponsorId": user_id}))
        
        if not team_members:
            return {"success": True, "data": []}
        
        # Batch fetch all users
        user_ids = [ObjectId(member["userId"]) for member in team_members]
        users_list = list(users_collection.find({"_id": {"$in": user_ids}}))
        users_map = {str(user["_id"]): user for user in users_list}
        
        # Batch fetch all plans
        plans_list = list(plans_collection.find({}))
        plans_map = {str(plan["_id"]): plan for plan in plans_list}
        plans_by_name = {plan["name"]: plan for plan in plans_list}
        
        result = []
        for member in team_members:
            user = users_map.get(member["userId"])
            if user:
                # Get plan name if exists
                plan_name = None
                if user.get("currentPlan"):
                    # Try ObjectId lookup first
                    plan = plans_map.get(user.get("currentPlan"))
                    if plan:
                        plan_name = plan.get("name")
                    else:
                        # Try by name
                        plan = plans_by_name.get(user.get("currentPlan"))
                        if plan:
                            plan_name = plan.get("name")
                        elif isinstance(user.get("currentPlan"), str) and len(user.get("currentPlan")) < 50:
                            plan_name = user.get("currentPlan")
                
                # Get user rank
                total_pv = user.get("totalPV", 0)
                user_rank = get_user_rank(total_pv)
                
                result.append({
                    "id": str(user["_id"]),
                    "name": user["name"],
                    "referralId": user["referralId"],
                    "mobile": user.get("mobile", ""),
                    "placement": member.get("placement"),
                    "currentPlan": plan_name,
                    "isActive": user.get("isActive", False),
                    "rank": user_rank,
                    "joinedAt": user.get("createdAt", get_ist_now()).isoformat()
                })
        
        return {
            "success": True,
            "data": serialize_doc(result)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ADMIN TEAM ROUTES ====================

@app.get("/api/admin/team/all")
async def get_all_teams(
    current_admin: dict = Depends(get_current_admin),
    search: Optional[str] = None,
    placement: Optional[str] = None
):
    """Get all teams (admin only)"""
    try:
        # Get all team relationships
        query = {}
        if placement and placement != "ALL":
            query["placement"] = placement.upper()
        
        teams = list(teams_collection.find(query))
        
        if not teams:
            return {
                "success": True,
                "data": {"members": [], "stats": {"totalMembers": 0, "leftCount": 0, "rightCount": 0}}
            }
        
        # Batch fetch all users and sponsors
        user_ids = [ObjectId(team["userId"]) for team in teams]
        sponsor_ids = [ObjectId(team["sponsorId"]) for team in teams]
        all_user_ids = list(set(user_ids + sponsor_ids))
        
        users_list = list(users_collection.find({"_id": {"$in": all_user_ids}}))
        users_map = {str(user["_id"]): user for user in users_list}
        
        # Batch fetch all plans
        plans_list = list(plans_collection.find({}))
        plans_map = {str(plan["_id"]): plan for plan in plans_list}
        plans_by_name = {plan["name"]: plan for plan in plans_list}
        
        result = []
        for team in teams:
            user = users_map.get(team["userId"])
            sponsor = users_map.get(team["sponsorId"])
            
            if user:
                # Get plan name if exists
                plan_name = None
                if user.get("currentPlan"):
                    # Try ObjectId lookup first
                    plan = plans_map.get(user.get("currentPlan"))
                    if plan:
                        plan_name = plan.get("name")
                    else:
                        # Try by name
                        plan = plans_by_name.get(user.get("currentPlan"))
                        if plan:
                            plan_name = plan.get("name")
                        elif isinstance(user.get("currentPlan"), str) and len(user.get("currentPlan")) < 50:
                            plan_name = user.get("currentPlan")
                
                # Get user rank
                total_pv = user.get("totalPV", 0)
                user_rank = get_user_rank(total_pv)
                
                member_data = {
                    "id": str(user["_id"]),
                    "name": user["name"],
                    "email": user.get("email", ""),
                    "mobile": user.get("mobile", ""),
                    "referralId": user["referralId"],
                    "placement": team.get("placement"),
                    "currentPlan": plan_name,
                    "isActive": user.get("isActive", False),
                    "rank": user_rank,
                    "joinedAt": user.get("createdAt", get_ist_now()).isoformat(),
                    "sponsorName": sponsor["name"] if sponsor else "N/A",
                    "sponsorId": sponsor["referralId"] if sponsor else "N/A"
                }
                
                # Apply search filter if provided
                if search:
                    search_lower = search.lower()
                    if (search_lower in member_data["name"].lower() or
                        search_lower in member_data["referralId"].lower() or
                        search_lower in member_data.get("email", "").lower()):
                        result.append(member_data)
                else:
                    result.append(member_data)
        
        # Calculate stats
        left_count = len([t for t in teams if t.get("placement") == "LEFT"])
        right_count = len([t for t in teams if t.get("placement") == "RIGHT"])
        
        return {
            "success": True,
            "data": {
                "members": serialize_doc(result),
                "stats": {
                    "totalMembers": len(result),
                    "leftMembers": left_count,
                    "rightMembers": right_count
                }
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/team/tree/{user_id}")
async def get_admin_team_tree(
    user_id: str,
    current_admin: dict = Depends(get_current_admin)
):
    """Get team tree for any user (admin only)"""
    try:
        def build_tree(parent_id: str, depth=0, max_depth=5) -> dict:
            if depth > max_depth:
                return None
                
            user = users_collection.find_one({"_id": ObjectId(parent_id)})
            if not user:
                return None
            
            # Get left and right children from teams collection
            left_child = teams_collection.find_one({
                "sponsorId": str(user["_id"]),
                "placement": "LEFT"
            })
            right_child = teams_collection.find_one({
                "sponsorId": str(user["_id"]),
                "placement": "RIGHT"
            })
            
            # Get plan name if exists
            plan_name = None
            if user.get("currentPlan"):
                try:
                    plan = plans_collection.find_one({"_id": ObjectId(user["currentPlan"])}, {"_id": 0})
                    if plan:
                        plan_name = plan.get("name")
                except Exception:
                    # If currentPlan is already a string (plan name), use it
                    plan_name = user.get("currentPlan") if isinstance(user.get("currentPlan"), str) and len(user.get("currentPlan")) < 50 else None
            
            node = {
                "id": str(user["_id"]),
                "name": user["name"],
                "referralId": user["referralId"],
                "placement": user.get("placement"),
                "currentPlan": plan_name,
                "isActive": user.get("isActive", False),
                "leftPV": user.get("leftPV", 0),
                "rightPV": user.get("rightPV", 0),
                "totalPV": user.get("totalPV", 0),
                "left": None,
                "right": None
            }
            
            if left_child:
                node["left"] = build_tree(left_child["userId"], depth + 1, max_depth)
            
            if right_child:
                node["right"] = build_tree(right_child["userId"], depth + 1, max_depth)
            
            return node
        
        tree = build_tree(user_id)
        
        return {
            "success": True,
            "data": serialize_doc(tree)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== PLANS ROUTES ====================

@app.get("/api/plans")
async def get_plans():
    """Get all active plans"""
    try:
        plans = list(plans_collection.find({"isActive": True}))
        return {
            "success": True,
            "data": serialize_doc(plans)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/plans/activate")
async def activate_plan(
    data: dict = Body(...),
    current_user: dict = Depends(get_current_active_user)
):
    """Activate a plan for user"""
    try:
        plan_id = data.get("planId")
        if not plan_id:
            raise HTTPException(status_code=400, detail="Plan ID required")
        
        # Get plan
        plan = plans_collection.find_one({"_id": ObjectId(plan_id)})
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")
        
        user_id = current_user["id"]
        user = users_collection.find_one({"_id": ObjectId(user_id)})
        
        # Get admin user for crediting plan activation amount
        admin_user = users_collection.find_one({"role": "admin"})
        admin_id = str(admin_user["_id"]) if admin_user else None
        
        # Update user's current plan
        users_collection.update_one(
            {"_id": ObjectId(user_id)},
            {
                "$set": {
                    "currentPlan": str(plan["_id"]),
                    "currentPlanName": plan["name"],
                    "dailyPVLimit": plan.get("dailyCapping", 500) // 25,  # Daily PV limit
                    "updatedAt": get_ist_now()
                }
            }
        )
        
        # Create PLAN_ACTIVATION transaction - this is ADMIN's REVENUE
        # Store with admin's userId so it shows in admin earnings
        transactions_collection.insert_one({
            "userId": admin_id if admin_id else user_id,  # Credit to admin
            "fromUserId": user_id,  # Track which user activated
            "type": "PLAN_ACTIVATION",
            "amount": plan["amount"],
            "description": f"{user.get('name', 'User')} activated {plan['name']} plan - ₹{plan['amount']}",
            "planName": plan["name"],
            "status": "COMPLETED",
            "createdAt": get_ist_now()
        })
        
        # Update admin wallet with plan activation amount (REVENUE)
        if admin_id:
            wallets_collection.update_one(
                {"userId": admin_id},
                {
                    "$inc": {
                        "balance": plan["amount"],
                        "totalEarnings": plan["amount"]
                    },
                    "$set": {"updatedAt": get_ist_now()}
                },
                upsert=True
            )
        
        # Distribute PV upward in the binary tree
        pv_amount = plan.get("pv", 0)
        if pv_amount > 0:
            distribute_pv_upward(user_id, pv_amount)
        
        # REFERRAL INCOME REMOVED - No longer giving referral income to sponsor
        # if current_user.get("sponsorId"):
        #     sponsor = users_collection.find_one({"referralId": current_user["sponsorId"]})
        #     if sponsor:
        #         sponsor_id = str(sponsor["_id"])
        #         
        #         # Update sponsor wallet
        #         wallets_collection.update_one(
        #             {"userId": sponsor_id},
        #             {
        #                 "$inc": {
        #                     "balance": plan["referralIncome"],
        #                     "totalEarnings": plan["referralIncome"]
        #                 },
        #                 "$set": {"updatedAt": get_ist_now()}
        #             }
        #         )
        #         
        #         # Create transaction for sponsor
        #         transactions_collection.insert_one({
        #             "userId": sponsor_id,
        #             "type": "REFERRAL_INCOME",
        #             "amount": plan["referralIncome"],
        #             "description": f"Referral income from {current_user['name']}",
        #             "status": "COMPLETED",
        #             "fromUser": user_id,
        #             "createdAt": get_ist_now()
        #         })
        
        return {
            "success": True,
            "message": "Plan activated successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/topup/request")
async def create_topup_request(
    data: dict = Body(...),
    current_user: dict = Depends(get_current_active_user)
):
    """Create a topup/plan upgrade request"""
    try:
        plan_id = data.get("planId")
        payment_method = data.get("paymentMethod", "Bank Transfer")
        transaction_details = data.get("transactionDetails", "")
        
        if not plan_id:
            raise HTTPException(status_code=400, detail="Plan ID required")
        
        if not transaction_details:
            raise HTTPException(status_code=400, detail="Transaction details required")
        
        # Get plan details
        plan = plans_collection.find_one({"_id": ObjectId(plan_id)})
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")
        
        user_id = current_user["id"]
        
        # Create topup request
        topup_request = {
            "userId": user_id,
            "planId": str(plan["_id"]),
            "amount": plan["amount"],
            "paymentMethod": payment_method,
            "transactionDetails": transaction_details,
            "status": "PENDING",
            "requestedAt": datetime.now(IST),
            "createdAt": datetime.now(IST)
        }
        
        result = topups_collection.insert_one(topup_request)
        
        return {
            "success": True,
            "message": "Topup request submitted successfully. Waiting for admin approval.",
            "data": {
                "requestId": str(result.inserted_id),
                "status": "PENDING"
            }
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ BINARY MLM PV DISTRIBUTION & MATCHING INCOME ============

def distribute_pv_upward(user_id: str, pv_amount: int):
    """
    Distribute PV upward in the binary tree
    PV flows completely to all sponsors based on placement
    """
    try:
        current_user = users_collection.find_one({"_id": ObjectId(user_id)})
        if not current_user:
            return
        
        # Get user's team record to find placement
        team_record = teams_collection.find_one({"userId": user_id})
        if not team_record or not team_record.get("sponsorId"):
            return  # No sponsor (admin user)
        
        placement = team_record.get("placement")  # LEFT or RIGHT
        sponsor_id = team_record["sponsorId"]
        
        # Travel up the tree
        while sponsor_id:
            sponsor = users_collection.find_one({"_id": ObjectId(sponsor_id)})
            if not sponsor:
                break
            
            # Add PV to sponsor's left or right leg based on placement
            update_field = "leftPV" if placement == "LEFT" else "rightPV"
            
            users_collection.update_one(
                {"_id": ObjectId(sponsor_id)},
                {
                    "$inc": {update_field: pv_amount},
                    "$set": {"updatedAt": get_ist_now()}
                }
            )
            
            # Note: Matching income will be calculated at end of day
            # Not calculated immediately to allow PV accumulation
            
            # Move up to next sponsor
            sponsor_team = teams_collection.find_one({"userId": sponsor_id})
            if not sponsor_team or not sponsor_team.get("sponsorId"):
                break
            
            # Get placement of current sponsor in their sponsor's tree
            placement = sponsor_team.get("placement")
            sponsor_id = sponsor_team["sponsorId"]
            
    except Exception as e:
        print(f"Error in PV distribution: {str(e)}")


def calculate_matching_income(user_id: str):
    """
    Calculate binary matching income based on left and right PV
    Formula: min(leftPV, rightPV) with daily capping
    Amount = todayPV × ₹25
    """
    try:
        user = users_collection.find_one({"_id": ObjectId(user_id)})
        if not user or not user.get("currentPlan"):
            return  # User must have an active plan
        
        # Get user's current PV
        left_pv = user.get("leftPV", 0)
        right_pv = user.get("rightPV", 0)
        
        # No matching possible if any side is 0
        if left_pv == 0 or right_pv == 0:
            return
        
        # Get plan details
        plan = plans_collection.find_one({"_id": ObjectId(user["currentPlan"])})
        if not plan:
            return
        
        daily_capping = plan.get("dailyCapping", 500)
        matching_income_rate = 25  # ₹25 per PV (as per your formula)
        
        # Calculate matching PV = min(leftPV, rightPV)
        matched_pv = min(left_pv, right_pv)
        
        # Check daily capping
        today_date = get_ist_now().replace(hour=0, minute=0, second=0, microsecond=0)
        last_matching_date = user.get("lastMatchingDate")
        
        # Reset daily PV if new day
        if not last_matching_date or last_matching_date.replace(hour=0, minute=0, second=0, microsecond=0) != today_date:
            daily_pv_used = 0
        else:
            daily_pv_used = user.get("dailyPVUsed", 0)
        
        # Calculate maximum PV allowed today (based on capping)
        max_pv_per_day = daily_capping // matching_income_rate  # 500 / 25 = 20 PV max per day
        remaining_pv_today = max_pv_per_day - daily_pv_used
        
        if remaining_pv_today <= 0:
            return  # Daily limit reached
        
        # Today's PV = min(matched_pv, remaining_pv_today)
        today_pv = min(matched_pv, remaining_pv_today)
        
        if today_pv <= 0:
            return
        
        # Calculate income
        income = today_pv * matching_income_rate
        
        # Update user's wallet
        wallets_collection.update_one(
            {"userId": user_id},
            {
                "$inc": {
                    "balance": income,
                    "totalEarnings": income
                },
                "$set": {"updatedAt": get_ist_now()}
            }
        )
        
        # Create transaction
        transactions_collection.insert_one({
            "userId": user_id,
            "type": "MATCHING_INCOME",
            "amount": income,
            "description": f"Binary matching income - {today_pv} PV @ ₹{matching_income_rate}/PV",
            "pv": today_pv,
            "status": "COMPLETED",
            "createdAt": get_ist_now()
        })
        
        # Flush matched PV from both sides
        users_collection.update_one(
            {"_id": ObjectId(user_id)},
            {
                "$inc": {
                    "leftPV": -today_pv,
                    "rightPV": -today_pv,
                    "totalPV": today_pv  # totalPV = lifetime PV earned
                },
                "$set": {
                    "lastMatchingDate": today_date,
                    "dailyPVUsed": daily_pv_used + today_pv,
                    "updatedAt": get_ist_now()
                }
            }
        )
        
        print(f"Matching income calculated for {user_id}: {income} (PV: {today_pv})")
        
    except Exception as e:
        print(f"Error in matching income calculation: {str(e)}")


def process_eod_matching_for_all_users():
    """
    Process EOD matching calculation for all active users
    This runs the matching calculation and carry forward logic
    """
    try:
        # Get all active users with a plan
        active_users = list(users_collection.find({
            "isActive": True,
            "currentPlan": {"$ne": None}
        }))
        
        processed_count = 0
        total_income = 0
        
        for user in active_users:
            try:
                user_id = str(user["_id"])
                left_pv = user.get("leftPV", 0)
                right_pv = user.get("rightPV", 0)
                
                # Only process if both sides have PV
                if left_pv > 0 and right_pv > 0:
                    # Get pre-calculation balance
                    wallet = wallets_collection.find_one({"userId": user_id})
                    pre_balance = wallet.get("balance", 0) if wallet else 0
                    
                    # Calculate matching income
                    calculate_matching_income(user_id)
                    
                    # Get post-calculation balance
                    wallet = wallets_collection.find_one({"userId": user_id})
                    post_balance = wallet.get("balance", 0) if wallet else 0
                    
                    income_earned = post_balance - pre_balance
                    if income_earned > 0:
                        total_income += income_earned
                        processed_count += 1
                        
            except Exception as user_error:
                print(f"Error processing user {user.get('referralId')}: {str(user_error)}")
                continue
        
        return {
            "processedUsers": processed_count,
            "totalIncomeDistributed": total_income,
            "timestamp": get_ist_now().isoformat()
        }
        
    except Exception as e:
        print(f"Error in EOD matching process: {str(e)}")
        return {"error": str(e)}


def process_carry_forward():
    """
    Process carry forward of unmatched PV to next day
    In binary MLM, unmatched PV (the difference) carries forward
    """
    try:
        # Get all active users
        active_users = list(users_collection.find({
            "isActive": True,
            "currentPlan": {"$ne": None}
        }))
        
        carried_forward_count = 0
        
        for user in active_users:
            try:
                user_id = str(user["_id"])
                left_pv = user.get("leftPV", 0)
                right_pv = user.get("rightPV", 0)
                
                # Calculate carry forward (weaker leg's remaining PV after matching)
                matched_pv = min(left_pv, right_pv)
                
                # After matching, subtract matched PV from both sides
                # Stronger leg retains the difference (carry forward)
                # This is already handled in calculate_matching_income
                
                # Reset daily PV used at EOD
                users_collection.update_one(
                    {"_id": user["_id"]},
                    {
                        "$set": {
                            "dailyPVUsed": 0,
                            "lastEODProcessed": get_ist_now(),
                            "updatedAt": get_ist_now()
                        }
                    }
                )
                carried_forward_count += 1
                
            except Exception as user_error:
                print(f"Error in carry forward for user {user.get('referralId')}: {str(user_error)}")
                continue
        
        return {"usersProcessed": carried_forward_count}
        
    except Exception as e:
        print(f"Error in carry forward process: {str(e)}")
        return {"error": str(e)}


# ==================== WALLET & TRANSACTIONS ====================

@app.get("/api/wallet/balance")
async def get_wallet_balance(current_user: dict = Depends(get_current_active_user)):
    """Get wallet balance"""
    try:
        wallet = wallets_collection.find_one({"userId": current_user["id"]})
        if not wallet:
            return {
                "success": True,
                "data": {
                    "balance": 0,
                    "totalEarnings": 0,
                    "totalWithdrawals": 0
                }
            }
        
        return {
            "success": True,
            "data": serialize_doc(wallet)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/wallet/transactions")
async def get_transactions(
    current_user: dict = Depends(get_current_active_user),
    limit: int = 50,
    skip: int = 0
):
    """Get user transactions (excluding PLAN_ACTIVATION)"""
    try:
        # Exclude PLAN_ACTIVATION transactions (admin income, not user income)
        query = {
            "userId": current_user["id"],
            "type": {"$ne": "PLAN_ACTIVATION"}
        }
        
        transactions = list(transactions_collection.find(query).sort("createdAt", DESCENDING).skip(skip).limit(limit))
        
        total = transactions_collection.count_documents(query)
        
        return {
            "success": True,
            "data": serialize_doc(transactions),
            "total": total,
            "limit": limit,
            "skip": skip
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== WITHDRAWAL ====================

@app.post("/api/withdrawal/request")
async def create_withdrawal_request(
    data: dict = Body(...),
    current_user: dict = Depends(get_current_active_user)
):
    """Create withdrawal request"""
    try:
        amount = data.get("amount")
        bank_details = data.get("bankDetails", {})
        
        if not amount or amount <= 0:
            raise HTTPException(status_code=400, detail="Invalid amount")
        
        # Get minimum withdraw limit from settings
        settings = settings_collection.find_one({})
        minimum_withdraw_limit = int(settings.get("minimumWithdrawLimit", 1000)) if settings else 1000
        
        if amount < minimum_withdraw_limit:
            raise HTTPException(status_code=400, detail=f"Minimum withdrawal amount is ₹{minimum_withdraw_limit}")
        
        # Check wallet balance
        wallet = wallets_collection.find_one({"userId": current_user["id"]})
        if not wallet or wallet.get("balance", 0) < amount:
            raise HTTPException(status_code=400, detail="Insufficient balance")
        
        # Create withdrawal request
        withdrawal = {
            "userId": current_user["id"],
            "amount": amount,
            "bankDetails": bank_details,
            "status": "PENDING",
            "requestedAt": get_ist_now(),
            "processedAt": None,
            "processedBy": None
        }
        
        result = withdrawals_collection.insert_one(withdrawal)
        
        # Deduct from balance (hold)
        wallets_collection.update_one(
            {"userId": current_user["id"]},
            {"$inc": {"balance": -amount}}
        )
        
        # Create transaction
        transactions_collection.insert_one({
            "userId": current_user["id"],
            "type": "WITHDRAWAL_REQUEST",
            "amount": -amount,
            "description": "Withdrawal request created",
            "status": "PENDING",
            "withdrawalId": str(result.inserted_id),
            "createdAt": get_ist_now()
        })
        
        return {
            "success": True,
            "message": "Withdrawal request created successfully",
            "withdrawalId": str(result.inserted_id)
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/withdrawal/history")
async def get_withdrawal_history(current_user: dict = Depends(get_current_active_user)):
    """Get withdrawal history"""
    try:
        withdrawals = list(withdrawals_collection.find(
            {"userId": current_user["id"]}
        ).sort("requestedAt", DESCENDING))
        
        return {
            "success": True,
            "data": serialize_doc(withdrawals)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== SETTINGS ROUTES ====================

@app.get("/api/settings/public")
async def get_public_settings():
    """Get public settings"""
    try:
        settings = settings_collection.find_one({})
        if not settings:
            # Return default settings
            return {
                "success": True,
                "data": {
                    "companyName": "VSV Unite",
                    "companyEmail": "info@vsvunite.com",
                    "companyPhone": "+91 9999999999",
                    "metaTitle": "VSV Unite - MLM Platform",
                    "metaDescription": "Join VSV Unite for transparent MLM opportunities"
                }
            }
        
        return {
            "success": True,
            "data": serialize_doc(settings)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/system/time")
async def get_system_time():
    """Get current system time (IST with offset)"""
    try:
        current_time = get_ist_now()
        eod_time = get_eod_time()
        offset = get_system_time_offset()
        
        return {
            "success": True,
            "data": {
                "currentTime": current_time.isoformat(),
                "currentTimeFormatted": current_time.strftime("%d-%m-%Y %I:%M:%S %p IST"),
                "eodTime": eod_time,
                "offsetMinutes": offset,
                "timezone": "Asia/Kolkata"
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/settings")
async def get_settings():
    """Get all settings (admin only)"""
    try:
        settings = settings_collection.find_one({})
        if not settings:
            return {"success": True, "data": {}}
        
        return {
            "success": True,
            "data": serialize_doc(settings)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/settings/general")
async def update_general_settings(data: dict = Body(...)):
    """Update general settings"""
    try:
        settings_collection.update_one(
            {},
            {"$set": {**data, "updatedAt": get_ist_now()}},
            upsert=True
        )
        return {"success": True, "message": "Settings updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/settings/seo")
async def update_seo_settings(data: dict = Body(...)):
    """Update SEO settings"""
    try:
        settings_collection.update_one(
            {},
            {"$set": {**data, "updatedAt": get_ist_now()}},
            upsert=True
        )
        return {"success": True, "message": "SEO settings updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/settings/hero")
async def update_hero_settings(data: dict = Body(...)):
    """Update hero settings"""
    try:
        settings_collection.update_one(
            {},
            {"$set": {**data, "updatedAt": get_ist_now()}},
            upsert=True
        )
        return {"success": True, "message": "Hero settings updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/settings/email-configuration")
async def get_email_config():
    """Get email configuration"""
    try:
        config = email_configs_collection.find_one({})
        if not config:
            return {"success": True, "emailConfig": None}
        
        config_data = serialize_doc(config)
        # Mask password
        if "smtpPassword" in config_data:
            config_data["smtpPassword"] = "****"
        
        return {"success": True, "emailConfig": config_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/settings/email-configuration")
async def update_email_config(data: dict = Body(...)):
    """Update email configuration"""
    try:
        email_configs_collection.update_one(
            {},
            {"$set": {**data, "updatedAt": get_ist_now()}},
            upsert=True
        )
        return {"success": True, "message": "Email configuration updated"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Ranks Management
@app.get("/api/settings/ranks")
async def get_ranks():
    """Get all ranks"""
    try:
        ranks = list(ranks_collection.find({}).sort("order", ASCENDING))
        return {"success": True, "data": serialize_doc(ranks)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/settings/ranks")
async def save_ranks(data: dict = Body(...)):
    """Save/update ranks"""
    try:
        ranks_data = data.get("ranks", [])
        
        # Clear existing ranks
        ranks_collection.delete_many({})
        
        # Insert new ranks
        if ranks_data:
            # Remove _id from new ranks if present
            for rank in ranks_data:
                rank.pop("_id", None)
            ranks_collection.insert_many(ranks_data)
        
        return {"success": True, "message": "Ranks updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/settings/ranks/{rank_id}")
async def delete_rank(rank_id: str):
    """Delete a rank"""
    try:
        result = ranks_collection.delete_one({"_id": ObjectId(rank_id)})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Rank not found")
        return {"success": True, "message": "Rank deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== ADMIN ROUTES ====================

@app.get("/api/admin/dashboard")
async def get_admin_dashboard(current_admin: dict = Depends(get_current_admin)):
    """Get admin dashboard statistics"""
    try:
        # Total users
        total_users = users_collection.count_documents({"role": "user"})
        active_users = users_collection.count_documents({"role": "user", "isActive": True})
        
        # Total earnings (sum of all wallets)
        pipeline = [
            {"$group": {
                "_id": None,
                "totalEarnings": {"$sum": "$totalEarnings"},
                "totalBalance": {"$sum": "$balance"},
                "totalWithdrawals": {"$sum": "$totalWithdrawals"}
            }}
        ]
        wallet_stats = list(wallets_collection.aggregate(pipeline))
        wallet_data = wallet_stats[0] if wallet_stats else {
            "totalEarnings": 0,
            "totalBalance": 0,
            "totalWithdrawals": 0
        }
        
        # Pending withdrawals
        pending_withdrawals = withdrawals_collection.count_documents({"status": "PENDING"})
        
        # Plan distribution - Use aggregation for efficiency
        pipeline = [
            {"$match": {"role": "user", "currentPlan": {"$ne": None, "$exists": True}}},
            {"$group": {"_id": "$currentPlan", "count": {"$sum": 1}}}
        ]
        plan_counts = {result["_id"]: result["count"] for result in users_collection.aggregate(pipeline)}
        
        # Get plan names
        plan_distribution = {}
        for plan in plans_collection.find({"isActive": True}):
            plan_id = str(plan["_id"])
            plan_name = plan["name"]
            # Check both ObjectId and name (for legacy data)
            count = plan_counts.get(plan_id, 0) + plan_counts.get(plan_name, 0)
            plan_distribution[plan_name] = count
        
        # Recent users
        recent_users = list(users_collection.find(
            {"role": "user"}
        ).sort("createdAt", DESCENDING).limit(5))
        
        return {
            "success": True,
            "data": {
                "users": {
                    "total": total_users,
                    "active": active_users,
                    "inactive": total_users - active_users
                },
                "earnings": wallet_data,
                "pendingWithdrawals": pending_withdrawals,
                "planDistribution": plan_distribution,
                "recentUsers": serialize_doc(recent_users)
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/earnings")
async def get_admin_earnings(current_admin: dict = Depends(get_current_admin)):
    """Get admin earnings - platform revenue and admin's personal earnings"""
    try:
        admin_id = current_admin["id"]
        
        # Get admin user data for PV info
        admin_user = users_collection.find_one({"_id": ObjectId(admin_id)})
        admin_left_pv = admin_user.get("leftPV", 0) if admin_user else 0
        admin_right_pv = admin_user.get("rightPV", 0) if admin_user else 0
        admin_total_pv = admin_user.get("totalPV", 0) if admin_user else 0
        
        # Get admin's wallet
        admin_wallet = wallets_collection.find_one({"userId": admin_id})
        admin_wallet_balance = admin_wallet.get("balance", 0) if admin_wallet else 0
        admin_total_earnings = admin_wallet.get("totalEarnings", 0) if admin_wallet else 0
        admin_total_withdrawals = admin_wallet.get("totalWithdrawals", 0) if admin_wallet else 0
        
        # ============ PLATFORM REVENUE (ALL PLAN ACTIVATIONS) ============
        all_activations = list(transactions_collection.find({
            "type": "PLAN_ACTIVATION"
        }).sort("createdAt", DESCENDING))
        
        total_platform_revenue = sum(txn.get("amount", 0) for txn in all_activations)
        
        # ============ ADMIN'S OWN EARNINGS ============
        # Admin's matching income (admin's personal binary matching)
        admin_matching_txns = list(transactions_collection.find({
            "userId": admin_id,
            "type": "MATCHING_INCOME"
        }).sort("createdAt", DESCENDING))
        admin_matching_income = sum(txn.get("amount", 0) for txn in admin_matching_txns)
        
        # Admin's referral income (if any - currently disabled)
        admin_referral_txns = list(transactions_collection.find({
            "userId": admin_id,
            "type": "REFERRAL_INCOME"
        }).sort("createdAt", DESCENDING))
        admin_referral_income = sum(txn.get("amount", 0) for txn in admin_referral_txns)
        
        # Admin's level income (if any)
        admin_level_txns = list(transactions_collection.find({
            "userId": admin_id,
            "type": "LEVEL_INCOME"
        }).sort("createdAt", DESCENDING))
        admin_level_income = sum(txn.get("amount", 0) for txn in admin_level_txns)
        
        # ============ TOTAL PAYOUTS TO ALL USERS ============
        all_matching_paid = list(transactions_collection.find({
            "type": "MATCHING_INCOME"
        }).sort("createdAt", DESCENDING))
        total_matching_paid = sum(txn.get("amount", 0) for txn in all_matching_paid)
        
        all_referral_paid = list(transactions_collection.find({
            "type": "REFERRAL_INCOME"
        }))
        total_referral_paid = sum(txn.get("amount", 0) for txn in all_referral_paid)
        
        all_level_paid = list(transactions_collection.find({
            "type": "LEVEL_INCOME"
        }))
        total_level_paid = sum(txn.get("amount", 0) for txn in all_level_paid)
        
        total_payouts = total_matching_paid + total_referral_paid + total_level_paid
        
        # Net Profit = Platform Revenue - Total Payouts
        net_profit = total_platform_revenue - total_payouts
        
        # Admin's total personal earnings
        admin_personal_earnings = admin_matching_income + admin_referral_income + admin_level_income
        
        # Income breakdown
        income_breakdown = {
            "PLAN_ACTIVATION": total_platform_revenue,
            "MATCHING_INCOME_PAID": total_matching_paid,
            "REFERRAL_INCOME_PAID": total_referral_paid,
            "LEVEL_INCOME_PAID": total_level_paid,
            "TOTAL_PAYOUTS": total_payouts,
            "NET_PROFIT": net_profit
        }
        
        # Admin's personal earnings breakdown
        admin_earnings_breakdown = {
            "MATCHING_INCOME": admin_matching_income,
            "REFERRAL_INCOME": admin_referral_income,
            "LEVEL_INCOME": admin_level_income,
            "TOTAL": admin_personal_earnings
        }
        
        # Plan activation breakdown by plan name
        income_by_plan = {}
        for txn in all_activations:
            desc = txn.get("description", "")
            for plan in ["Basic", "Standard", "Advanced", "Premium"]:
                if plan in desc:
                    income_by_plan[plan] = income_by_plan.get(plan, 0) + txn.get("amount", 0)
                    break
        
        # Today's calculations - handle timezone-naive dates safely
        now = get_ist_now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
        today_revenue = 0
        today_matching_paid_amount = 0
        month_revenue = 0
        
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        for txn in all_activations:
            created_at = txn.get("createdAt")
            if created_at:
                try:
                    # Make timezone-naive for comparison if needed
                    if hasattr(created_at, 'tzinfo') and created_at.tzinfo is not None:
                        txn_date = created_at
                    else:
                        txn_date = IST.localize(created_at) if created_at else None
                    
                    if txn_date:
                        if txn_date >= today_start:
                            today_revenue += txn.get("amount", 0)
                        if txn_date >= month_start:
                            month_revenue += txn.get("amount", 0)
                except:
                    pass
        
        for txn in all_matching_paid:
            created_at = txn.get("createdAt")
            if created_at:
                try:
                    if hasattr(created_at, 'tzinfo') and created_at.tzinfo is not None:
                        txn_date = created_at
                    else:
                        txn_date = IST.localize(created_at) if created_at else None
                    
                    if txn_date and txn_date >= today_start:
                        today_matching_paid_amount += txn.get("amount", 0)
                except:
                    pass
        
        # Recent transactions (all types)
        recent_transactions = []
        
        # Get all income transactions
        all_income_txns = list(transactions_collection.find({
            "type": {"$in": ["PLAN_ACTIVATION", "MATCHING_INCOME", "REFERRAL_INCOME", "LEVEL_INCOME"]}
        }).sort("createdAt", DESCENDING).limit(50))
        
        for txn in all_income_txns:
            user = None
            user_id = txn.get("userId")
            if user_id:
                try:
                    user = users_collection.find_one({"_id": ObjectId(user_id)})
                except:
                    pass
            
            # Check if this is admin's own transaction
            is_admin_txn = user_id == admin_id
            
            recent_transactions.append({
                "id": str(txn["_id"]),
                "type": txn.get("type"),
                "userName": user.get("name") if user else "System",
                "userReferralId": user.get("referralId") if user else "-",
                "amount": txn.get("amount"),
                "description": txn.get("description"),
                "createdAt": txn.get("createdAt"),
                "isAdminTransaction": is_admin_txn
            })
        
        return {
            "success": True,
            "data": {
                # Platform Stats
                "totalRevenue": total_platform_revenue,
                "totalPayouts": total_payouts,
                "netProfit": net_profit,
                
                # Payout breakdown
                "totalMatchingPaid": total_matching_paid,
                "totalReferralPaid": total_referral_paid,
                "totalLevelPaid": total_level_paid,
                
                # Today's stats
                "todayRevenue": today_revenue,
                "todayMatchingPaid": today_matching_paid_amount,
                "monthRevenue": month_revenue,
                
                # Admin's personal earnings
                "adminWallet": {
                    "balance": admin_wallet_balance,
                    "totalEarnings": admin_total_earnings,
                    "totalWithdrawals": admin_total_withdrawals
                },
                "adminEarnings": admin_earnings_breakdown,
                
                # Admin's PV stats
                "adminPV": {
                    "leftPV": admin_left_pv,
                    "rightPV": admin_right_pv,
                    "totalPV": admin_total_pv,
                    "matchablePV": min(admin_left_pv, admin_right_pv)
                },
                
                # Breakdowns
                "incomeBreakdown": income_breakdown,
                "totalActivations": len(all_activations),
                "incomeByPlan": income_by_plan,
                
                # Transactions
                "recentTransactions": serialize_doc(recent_transactions),
                "allTransactions": serialize_doc(all_activations)
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/users")
async def get_all_users(
    current_admin: dict = Depends(get_current_admin),
    limit: int = 50,
    skip: int = 0,
    search: Optional[str] = None
):
    """Get all users (admin only)"""
    try:
        # Include all users (both admin and user roles)
        query = {}
        
        if search:
            query["$or"] = [
                {"name": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}},
                {"referralId": {"$regex": search, "$options": "i"}},
                {"mobile": {"$regex": search, "$options": "i"}}
            ]
        
        users = list(users_collection.find(query).skip(skip).limit(limit))
        total = users_collection.count_documents(query)
        
        # Batch fetch all plans
        plans_list = list(plans_collection.find({}))
        plans_map = {str(plan["_id"]): plan for plan in plans_list}
        plans_by_name = {plan["name"]: plan for plan in plans_list}
        
        # Batch fetch placement information from teams collection
        user_ids = [str(user["_id"]) for user in users]
        teams_data = list(teams_collection.find({"userId": {"$in": user_ids}}))
        teams_map = {team["userId"]: team for team in teams_data}
        
        # Remove passwords and convert plan IDs to names
        for user in users:
            user.pop("password", None)
            
            # Add placement from teams collection
            user_id = str(user["_id"])
            team_data = teams_map.get(user_id)
            if team_data:
                user["placement"] = team_data.get("placement")
            else:
                user["placement"] = None
            
            # Convert currentPlan ObjectId to plan name
            if user.get("currentPlan"):
                # Try ObjectId lookup
                plan = plans_map.get(user.get("currentPlan"))
                if plan:
                    user["currentPlan"] = plan.get("name")
                else:
                    # Try by name
                    plan = plans_by_name.get(user.get("currentPlan"))
                    if plan:
                        user["currentPlan"] = plan.get("name")
                    elif isinstance(user.get("currentPlan"), str) and len(user.get("currentPlan")) < 50:
                        # Keep as is if it's a valid string
                        pass
                    else:
                        user["currentPlan"] = None
        
        return {
            "success": True,
            "data": serialize_doc(users),
            "total": total,
            "limit": limit,
            "skip": skip
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/admin/users/{user_id}/status")
async def update_user_status(
    user_id: str,
    data: dict = Body(...),
    current_admin: dict = Depends(get_current_admin)
):
    """Update user status (activate/deactivate)"""
    try:
        is_active = data.get("isActive")
        if is_active is None:
            raise HTTPException(status_code=400, detail="isActive field required")
        
        users_collection.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": {"isActive": is_active, "updatedAt": get_ist_now()}}
        )
        
        return {
            "success": True,
            "message": f"User {'activated' if is_active else 'deactivated'} successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/admin/users/{user_id}")
async def update_user(
    user_id: str,
    data: dict = Body(...),
    current_admin: dict = Depends(get_current_admin)
):
    """Update user information (admin only)"""
    try:
        user = users_collection.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        update_data = {}
        if data.get("name"):
            update_data["name"] = data["name"]
        if data.get("email"):
            # Check if email is already taken by another user
            existing = users_collection.find_one({"email": data["email"], "_id": {"$ne": ObjectId(user_id)}})
            if existing:
                raise HTTPException(status_code=400, detail="Email already in use")
            update_data["email"] = data["email"]
        if data.get("mobile"):
            update_data["mobile"] = data["mobile"]
        if "currentPlan" in data:
            # Handle plan assignment/change
            if data["currentPlan"]:
                # Find plan by name
                plan = plans_collection.find_one({"name": data["currentPlan"]})
                if plan:
                    update_data["currentPlan"] = str(plan["_id"])
                else:
                    update_data["currentPlan"] = data["currentPlan"]
            else:
                update_data["currentPlan"] = None
        
        if update_data:
            update_data["updatedAt"] = get_ist_now()
            users_collection.update_one(
                {"_id": ObjectId(user_id)},
                {"$set": update_data}
            )
        
        return {
            "success": True,
            "message": "User updated successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/admin/users/{user_id}/reset-password")
async def reset_user_password(
    user_id: str,
    data: dict = Body(...),
    current_admin: dict = Depends(get_current_admin)
):
    """Reset user password (admin only)"""
    try:
        user = users_collection.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        new_password = data.get("newPassword")
        if not new_password or len(new_password) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
        
        hashed_password = hash_password(new_password)
        users_collection.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": {"password": hashed_password, "updatedAt": get_ist_now()}}
        )
        
        return {
            "success": True,
            "message": "Password reset successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/admin/users/{user_id}")
async def delete_user(
    user_id: str,
    current_admin: dict = Depends(get_current_admin)
):
    """Delete user (admin only)"""
    try:
        user = users_collection.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Delete user's wallet
        wallets_collection.delete_one({"userId": user_id})
        
        # Delete user's transactions
        transactions_collection.delete_many({"userId": user_id})
        
        # Delete user's team entries
        teams_collection.delete_many({"userId": user_id})
        
        # Delete user's withdrawals
        withdrawals_collection.delete_many({"userId": user_id})
        
        # Delete user
        users_collection.delete_one({"_id": ObjectId(user_id)})
        
        return {
            "success": True,
            "message": "User deleted successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/admin/withdrawals")
async def get_all_withdrawals(
    current_admin: dict = Depends(get_current_admin),
    status: Optional[str] = None
):
    """Get all withdrawal requests"""
    try:
        query = {}
        if status:
            query["status"] = status.upper()
        
        withdrawals = list(withdrawals_collection.find(query).sort("requestedAt", DESCENDING))
        
        if not withdrawals:
            return {"success": True, "data": []}
        
        # Batch fetch all users
        user_ids = [ObjectId(w["userId"]) for w in withdrawals]
        users_list = list(users_collection.find({"_id": {"$in": user_ids}}))
        users_map = {str(user["_id"]): user for user in users_list}
        
        # Add user details
        result = []
        for withdrawal in withdrawals:
            withdrawal_data = serialize_doc(withdrawal)
            user = users_map.get(withdrawal["userId"])
            if user:
                withdrawal_data["userName"] = user["name"]
                withdrawal_data["userEmail"] = user.get("email", "")
                withdrawal_data["userMobile"] = user.get("mobile", "")
            result.append(withdrawal_data)
        
        return {
            "success": True,
            "data": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/admin/withdrawals/{withdrawal_id}/approve")
async def approve_withdrawal(
    withdrawal_id: str,
    current_admin: dict = Depends(get_current_admin)
):
    """Approve withdrawal request"""
    try:
        withdrawal = withdrawals_collection.find_one({"_id": ObjectId(withdrawal_id)})
        if not withdrawal:
            raise HTTPException(status_code=404, detail="Withdrawal not found")
        
        if withdrawal["status"] != "PENDING":
            raise HTTPException(status_code=400, detail="Withdrawal already processed")
        
        # Update withdrawal status
        withdrawals_collection.update_one(
            {"_id": ObjectId(withdrawal_id)},
            {
                "$set": {
                    "status": "APPROVED",
                    "processedAt": get_ist_now(),
                    "processedBy": current_admin["id"]
                }
            }
        )
        
        # Update wallet
        wallets_collection.update_one(
            {"userId": withdrawal["userId"]},
            {"$inc": {"totalWithdrawals": withdrawal["amount"]}}
        )
        
        # Update transaction
        transactions_collection.update_one(
            {"withdrawalId": withdrawal_id},
            {"$set": {"status": "COMPLETED"}}
        )
        
        return {
            "success": True,
            "message": "Withdrawal approved successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/admin/withdrawals/{withdrawal_id}/reject")
async def reject_withdrawal(
    withdrawal_id: str,
    data: dict = Body(...),
    current_admin: dict = Depends(get_current_admin)
):
    """Reject withdrawal request"""
    try:
        withdrawal = withdrawals_collection.find_one({"_id": ObjectId(withdrawal_id)})
        if not withdrawal:
            raise HTTPException(status_code=404, detail="Withdrawal not found")
        
        if withdrawal["status"] != "PENDING":
            raise HTTPException(status_code=400, detail="Withdrawal already processed")
        
        reason = data.get("reason", "No reason provided")
        
        # Update withdrawal status
        withdrawals_collection.update_one(
            {"_id": ObjectId(withdrawal_id)},
            {
                "$set": {
                    "status": "REJECTED",
                    "rejectionReason": reason,
                    "processedAt": get_ist_now(),
                    "processedBy": current_admin["id"]
                }
            }
        )
        
        # Return amount to wallet
        wallets_collection.update_one(
            {"userId": withdrawal["userId"]},
            {"$inc": {"balance": withdrawal["amount"]}}
        )
        
        # Update transaction
        transactions_collection.update_one(
            {"withdrawalId": withdrawal_id},
            {"$set": {"status": "REJECTED"}}
        )
        
        return {
            "success": True,
            "message": "Withdrawal rejected successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/plans")
async def get_admin_plans(current_admin: dict = Depends(get_current_admin)):
    """Get all plans (admin)"""
    try:
        plans = list(plans_collection.find({}))
        return {
            "success": True,
            "data": serialize_doc(plans)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/admin/plans")
async def create_plan(
    data: dict = Body(...),
    current_admin: dict = Depends(get_current_admin)
):
    """Create new plan"""
    try:
        plan_data = {
            **data,
            "isActive": data.get("isActive", True),
            "createdAt": get_ist_now(),
            "updatedAt": get_ist_now()
        }
        
        result = plans_collection.insert_one(plan_data)
        
        return {
            "success": True,
            "message": "Plan created successfully",
            "planId": str(result.inserted_id)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/admin/plans/{plan_id}")
async def update_plan(
    plan_id: str,
    data: dict = Body(...),
    current_admin: dict = Depends(get_current_admin)
):
    """Update plan"""
    try:
        data["updatedAt"] = get_ist_now()
        
        plans_collection.update_one(
            {"_id": ObjectId(plan_id)},
            {"$set": data}
        )
        
        return {
            "success": True,
            "message": "Plan updated successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/admin/plans/{plan_id}")
async def delete_plan(
    plan_id: str,
    current_admin: dict = Depends(get_current_admin)
):
    """Delete plan (admin only)"""
    try:
        plan = plans_collection.find_one({"_id": ObjectId(plan_id)})
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")
        
        # Check if any users have this plan
        users_with_plan = users_collection.count_documents({"currentPlanId": plan_id})
        if users_with_plan > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot delete plan. {users_with_plan} users are currently on this plan"
            )
        
        plans_collection.delete_one({"_id": ObjectId(plan_id)})
        
        return {
            "success": True,
            "message": "Plan deleted successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== HEALTH CHECK ====================

@app.get("/")
async def root():
    return {
        "message": "VSV Unite MLM API",
        "version": "1.0.0",
        "status": "running"
    }

# ============ TOPUP/PLAN ACTIVATION MANAGEMENT (ADMIN) ============

class TopupRequest(BaseModel):
    userId: str
    planId: str
    amount: float
    paymentMode: str
    transactionId: str
    paymentProof: Optional[str] = None

@app.get("/api/admin/topups")
async def get_all_topups(
    current_admin: dict = Depends(get_current_admin),
    status: Optional[str] = None
):
    """Get all topup/plan activation requests"""
    try:
        query = {}
        if status:
            query["status"] = status.upper()
        
        topups = list(topups_collection.find(query).sort("requestedAt", DESCENDING))
        
        if not topups:
            return {"success": True, "data": []}
        
        # Batch fetch users
        user_ids = [ObjectId(t["userId"]) for t in topups if t.get("userId")]
        users_list = list(users_collection.find({"_id": {"$in": user_ids}})) if user_ids else []
        users_map = {str(user["_id"]): user for user in users_list}
        
        # Batch fetch plans
        plan_ids = [ObjectId(t["planId"]) for t in topups if t.get("planId")]
        plans_list = list(plans_collection.find({"_id": {"$in": plan_ids}})) if plan_ids else []
        plans_map = {str(plan["_id"]): plan for plan in plans_list}
        
        # Enrich with user and plan details
        for topup in topups:
            if topup.get("userId"):
                user = users_map.get(topup["userId"])
                if user:
                    topup["userName"] = user.get("name")
                    topup["userEmail"] = user.get("email")
                    topup["referralId"] = user.get("referralId")
            
            if topup.get("planId"):
                plan = plans_map.get(topup["planId"])
                if plan:
                    topup["planName"] = plan.get("name")
        
        return {
            "success": True,
            "data": serialize_doc(topups)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/admin/topups/{topup_id}/approve")
async def approve_topup(
    topup_id: str,
    current_admin: dict = Depends(get_current_admin)
):
    """Approve a topup/plan activation request"""
    try:
        topup = topups_collection.find_one({"_id": ObjectId(topup_id)})
        if not topup:
            raise HTTPException(status_code=404, detail="Topup request not found")
        
        if topup["status"] != "PENDING":
            raise HTTPException(status_code=400, detail="Only pending requests can be approved")
        
        user_id = topup["userId"]
        plan_id = topup["planId"]
        
        # Get plan details
        plan = plans_collection.find_one({"_id": ObjectId(plan_id)})
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")
        
        # Update user's current plan
        users_collection.update_one(
            {"_id": ObjectId(user_id)},
            {
                "$set": {
                    "currentPlan": str(plan["_id"]),
                    "currentPlanName": plan["name"],
                    "dailyPVLimit": plan.get("dailyCapping", 500) // 25,
                    "updatedAt": get_ist_now()
                }
            }
        )
        
        # Create transaction for user
        transactions_collection.insert_one({
            "userId": user_id,
            "type": "PLAN_ACTIVATION",
            "amount": plan["amount"],
            "description": f"Activated {plan['name']} plan",
            "status": "COMPLETED",
            "createdAt": get_ist_now()
        })
        
        # Distribute PV upward in the binary tree
        pv_amount = plan.get("pv", 0)
        if pv_amount > 0:
            distribute_pv_upward(user_id, pv_amount)
        
        # Update topup status
        topups_collection.update_one(
            {"_id": ObjectId(topup_id)},
            {
                "$set": {
                    "status": "APPROVED",
                    "approvedAt": datetime.now(IST),
                    "approvedBy": current_admin["id"]
                }
            }
        )
        
        # REFERRAL INCOME REMOVED - No longer giving referral income to sponsor
        # user = users_collection.find_one({"_id": ObjectId(user_id)})
        # if user and user.get("sponsorId"):
        #     sponsor = users_collection.find_one({"referralId": user["sponsorId"]})
        #     if sponsor:
        #         sponsor_id = str(sponsor["_id"])
        #         referral_income = plan.get("referralIncome", 0)
        #         
        #         # Update sponsor wallet
        #         wallets_collection.update_one(
        #             {"userId": sponsor_id},
        #             {
        #                 "$inc": {
        #                     "balance": referral_income,
        #                     "totalEarnings": referral_income
        #                 },
        #                 "$set": {"updatedAt": datetime.now(IST)}
        #             }
        #         )
        #         
        #         # Create transaction for sponsor
        #         transactions_collection.insert_one({
        #             "userId": sponsor_id,
        #             "type": "REFERRAL_INCOME",
        #             "amount": referral_income,
        #             "description": f"Referral income from {user['name']} plan activation",
        #             "status": "COMPLETED",
        #             "fromUser": user_id,
        #             "createdAt": datetime.now(IST)
        #         })
        
        return {
            "success": True,
            "message": "Topup approved successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/admin/topups/{topup_id}/reject")
async def reject_topup(
    topup_id: str,
    data: dict = Body(...),
    current_admin: dict = Depends(get_current_admin)
):
    """Reject a topup/plan activation request"""
    try:
        topup = topups_collection.find_one({"_id": ObjectId(topup_id)})
        if not topup:
            raise HTTPException(status_code=404, detail="Topup request not found")
        
        if topup["status"] != "PENDING":
            raise HTTPException(status_code=400, detail="Only pending requests can be rejected")
        
        reason = data.get("reason", "Rejected by admin")
        
        topups_collection.update_one(
            {"_id": ObjectId(topup_id)},
            {
                "$set": {
                    "status": "REJECTED",
                    "rejectedAt": get_ist_now(),
                    "rejectedBy": current_admin["id"],
                    "rejectionReason": reason
                }
            }
        )
        
        return {
            "success": True,
            "message": "Topup rejected successfully"
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ REPORTS & ANALYTICS (ADMIN) ============

@app.get("/api/admin/reports/dashboard")
async def get_dashboard_reports(
    current_admin: dict = Depends(get_current_admin)
):
    """Get dashboard analytics and reports"""
    try:
        # Total users count (including all users - admin + regular users)
        total_users = users_collection.count_documents({})
        active_users = users_collection.count_documents({"isActive": True})
        inactive_users = users_collection.count_documents({"isActive": False})
        
        # Users with plans
        with_plans = users_collection.count_documents({
            "currentPlan": {"$exists": True, "$ne": None, "$ne": ""}
        })
        
        # Total earnings (sum of all credit transactions)
        total_earnings_pipeline = [
            {"$match": {"amount": {"$gt": 0}}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
        ]
        total_earnings_result = list(transactions_collection.aggregate(total_earnings_pipeline))
        total_earnings = total_earnings_result[0]["total"] if total_earnings_result else 0
        
        # Total withdrawals
        total_withdrawals_pipeline = [
            {"$match": {"status": "APPROVED"}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
        ]
        total_withdrawals_result = list(withdrawals_collection.aggregate(total_withdrawals_pipeline))
        total_withdrawals = total_withdrawals_result[0]["total"] if total_withdrawals_result else 0
        
        # Pending withdrawals
        pending_withdrawals = withdrawals_collection.count_documents({"status": "PENDING"})
        pending_withdrawals_amount_pipeline = [
            {"$match": {"status": "PENDING"}},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
        ]
        pending_withdrawals_amount_result = list(withdrawals_collection.aggregate(pending_withdrawals_amount_pipeline))
        pending_withdrawals_amount = pending_withdrawals_amount_result[0]["total"] if pending_withdrawals_amount_result else 0
        
        # Plan distribution
        plan_distribution = {}
        for plan in plans_collection.find():
            # Count users with this plan (handle both ObjectId and string formats)
            plan_id_str = str(plan["_id"])
            count = users_collection.count_documents({
                "$or": [
                    {"currentPlan": plan_id_str},
                    {"currentPlan": plan["_id"]},
                    {"currentPlan": plan["name"]}
                ]
            })
            plan_distribution[plan["name"]] = count
        
        # Recent registrations (last 7 days)
        seven_days_ago = get_ist_now() - timedelta(days=7)
        recent_registrations = users_collection.count_documents({
            "role": "user",
            "createdAt": {"$gte": seven_days_ago}
        })
        
        # Daily business report (last 7 days)
        daily_reports = []
        for i in range(6, -1, -1):
            day_start = get_ist_now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=i)
            day_end = day_start + timedelta(days=1)
            
            # New users on this day
            new_users = users_collection.count_documents({
                "role": "user",
                "createdAt": {"$gte": day_start, "$lt": day_end}
            })
            
            # Topups on this day
            topups_pipeline = [
                {"$match": {
                    "status": "APPROVED",
                    "approvedAt": {"$gte": day_start, "$lt": day_end}
                }},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ]
            topups_result = list(topups_collection.aggregate(topups_pipeline))
            topups_amount = topups_result[0]["total"] if topups_result else 0
            
            # Payouts on this day
            payouts_pipeline = [
                {"$match": {
                    "status": "APPROVED",
                    "approvedAt": {"$gte": day_start, "$lt": day_end}
                }},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ]
            payouts_result = list(withdrawals_collection.aggregate(payouts_pipeline))
            payouts_amount = payouts_result[0]["total"] if payouts_result else 0
            
            daily_reports.append({
                "date": day_start.strftime("%Y-%m-%d"),
                "newUsers": new_users,
                "topups": topups_amount,
                "payouts": payouts_amount,
                "netBusiness": topups_amount - payouts_amount
            })
        
        # Income breakdown
        income_types = {}
        for income_type in ["REFERRAL_INCOME", "MATCHING_INCOME", "LEVEL_INCOME"]:
            pipeline = [
                {"$match": {"type": income_type}},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ]
            result = list(transactions_collection.aggregate(pipeline))
            income_types[income_type] = result[0]["total"] if result else 0
        
        # Recent users
        recent_users = list(users_collection.find(
            {"role": "user"}
        ).sort("createdAt", DESCENDING).limit(5))
        
        return {
            "success": True,
            "data": {
                "overview": {
                    "totalUsers": total_users,
                    "activeUsers": active_users,
                    "inactiveUsers": inactive_users,
                    "withPlans": with_plans,
                    "totalEarnings": total_earnings,
                    "totalWithdrawals": total_withdrawals,
                    "pendingWithdrawals": pending_withdrawals,
                    "pendingWithdrawalsAmount": pending_withdrawals_amount,
                    "recentRegistrations": recent_registrations
                },
                "planDistribution": plan_distribution,
                "dailyReports": daily_reports,
                "incomeBreakdown": income_types,
                "recentUsers": serialize_doc(recent_users)
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ DOWNLOADABLE REPORTS ENDPOINTS ============

# USER REPORTS

@app.get("/api/admin/reports/users/all")
async def get_all_users_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get all members report with optional date filter"""
    try:
        start, end = parse_date_range(start_date, end_date)
        
        query = {"role": "user"}
        if start or end:
            query["createdAt"] = {}
            if start:
                query["createdAt"]["$gte"] = start
            if end:
                query["createdAt"]["$lte"] = end
        
        users = list(users_collection.find(query, {"password": 0}))
        
        if not users:
            return {"success": True, "data": [], "total": 0}
        
        # Batch fetch plans
        plans_list = list(plans_collection.find({}))
        plans_map = {str(plan["_id"]): plan for plan in plans_list}
        plans_by_name = {plan["name"]: plan for plan in plans_list}
        
        # Batch fetch wallets
        user_ids = [str(user["_id"]) for user in users]
        wallets_list = list(wallets_collection.find({"userId": {"$in": user_ids}}))
        wallets_map = {wallet["userId"]: wallet for wallet in wallets_list}
        
        # Format data
        report_data = []
        for user in users:
            plan_name = "No Plan"
            if user.get("currentPlan"):
                # Try ObjectId lookup
                plan = plans_map.get(user.get("currentPlan"))
                if plan:
                    plan_name = plan.get("name", "No Plan")
                else:
                    # Try by name
                    plan = plans_by_name.get(user.get("currentPlan"))
                    if plan:
                        plan_name = plan.get("name", "No Plan")
                    elif isinstance(user.get("currentPlan"), str):
                        plan_name = user.get("currentPlan")
            
            wallet = wallets_map.get(str(user["_id"]))
            balance = wallet.get("balance", 0) if wallet else 0
            
            report_data.append({
                "Referral ID": user.get("referralId", ""),
                "Name": user.get("name", ""),
                "Email": user.get("email", ""),
                "Mobile": user.get("mobile", ""),
                "Sponsor ID": user.get("sponsorId", ""),
                "Current Plan": plan_name,
                "Status": "Active" if user.get("isActive", False) else "Inactive",
                "Wallet Balance": f"₹{balance}",
                "Joined Date": user.get("createdAt", datetime.now()).strftime("%d-%m-%Y") if user.get("createdAt") else ""
            })
        
        if format == "excel":
            headers = ["Referral ID", "Name", "Email", "Mobile", "Sponsor ID", "Current Plan", "Status", "Wallet Balance", "Joined Date"]
            output = generate_excel_report(report_data, headers, "All Members Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=all_members_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Referral ID", "Name", "Email", "Current Plan", "Status", "Balance", "Joined"]
            pdf_data = []
            for item in report_data:
                pdf_data.append({
                    "Referral ID": item["Referral ID"],
                    "Name": item["Name"],
                    "Email": item["Email"],
                    "Current Plan": item["Current Plan"],
                    "Status": item["Status"],
                    "Balance": item["Wallet Balance"],
                    "Joined": item["Joined Date"]
                })
            output = generate_pdf_report(pdf_data, headers, "All Members Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=all_members_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {"success": True, "data": report_data, "total": len(report_data)}
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/users/active-inactive")
async def get_active_inactive_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get active/inactive users breakdown"""
    try:
        start, end = parse_date_range(start_date, end_date)
        
        base_query = {"role": "user"}
        if start or end:
            base_query["createdAt"] = {}
            if start:
                base_query["createdAt"]["$gte"] = start
            if end:
                base_query["createdAt"]["$lte"] = end
        
        # Get active users
        active_query = {**base_query, "isActive": True}
        active_users = list(users_collection.find(active_query, {"password": 0}))
        
        # Get inactive users
        inactive_query = {**base_query, "isActive": False}
        inactive_users = list(users_collection.find(inactive_query, {"password": 0}))
        
        report_data = []
        
        for user in active_users + inactive_users:
            report_data.append({
                "Referral ID": user.get("referralId", ""),
                "Name": user.get("name", ""),
                "Email": user.get("email", ""),
                "Status": "Active" if user.get("isActive", False) else "Inactive",
                "Joined Date": user.get("createdAt", datetime.now()).strftime("%d-%m-%Y") if user.get("createdAt") else ""
            })
        
        if format == "excel":
            headers = ["Referral ID", "Name", "Email", "Status", "Joined Date"]
            output = generate_excel_report(report_data, headers, "Active/Inactive Users Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=active_inactive_users_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Referral ID", "Name", "Email", "Status", "Joined Date"]
            output = generate_pdf_report(report_data, headers, "Active/Inactive Users Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=active_inactive_users_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {
                "success": True,
                "data": report_data,
                "summary": {
                    "total": len(report_data),
                    "active": len(active_users),
                    "inactive": len(inactive_users)
                }
            }
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/users/by-plan")
async def get_users_by_plan_report(
    plan_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get users by plan type"""
    try:
        start, end = parse_date_range(start_date, end_date)
        
        query = {"role": "user"}
        if start or end:
            query["createdAt"] = {}
            if start:
                query["createdAt"]["$gte"] = start
            if end:
                query["createdAt"]["$lte"] = end
        
        if plan_id and plan_id != "all":
            query["$or"] = [
                {"currentPlan": plan_id},
                {"currentPlan": ObjectId(plan_id) if len(plan_id) == 24 else plan_id}
            ]
        
        users = list(users_collection.find(query, {"password": 0}))
        
        report_data = []
        for user in users:
            plan_name = "No Plan"
            if user.get("currentPlan"):
                try:
                    plan = plans_collection.find_one({"_id": ObjectId(user["currentPlan"])})
                    if plan:
                        plan_name = plan.get("name", "No Plan")
                except:
                    pass
            
            report_data.append({
                "Referral ID": user.get("referralId", ""),
                "Name": user.get("name", ""),
                "Email": user.get("email", ""),
                "Plan": plan_name,
                "Status": "Active" if user.get("isActive", False) else "Inactive",
                "Joined Date": user.get("createdAt", datetime.now()).strftime("%d-%m-%Y") if user.get("createdAt") else ""
            })
        
        if format == "excel":
            headers = ["Referral ID", "Name", "Email", "Plan", "Status", "Joined Date"]
            output = generate_excel_report(report_data, headers, "Users by Plan Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=users_by_plan_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Referral ID", "Name", "Email", "Plan", "Status", "Joined Date"]
            output = generate_pdf_report(report_data, headers, "Users by Plan Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=users_by_plan_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {"success": True, "data": report_data, "total": len(report_data)}
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# FINANCIAL REPORTS

@app.get("/api/admin/reports/financial/earnings")
async def get_earnings_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get earnings summary report"""
    try:
        start, end = parse_date_range(start_date, end_date)
        
        query = {"amount": {"$gt": 0}}
        if start or end:
            query["createdAt"] = {}
            if start:
                query["createdAt"]["$gte"] = start
            if end:
                query["createdAt"]["$lte"] = end
        
        transactions = list(transactions_collection.find(query))
        
        report_data = []
        for txn in transactions:
            user = users_collection.find_one({"_id": ObjectId(txn.get("userId"))})
            report_data.append({
                "Date": txn.get("createdAt", datetime.now()).strftime("%d-%m-%Y %I:%M %p") if txn.get("createdAt") else "",
                "User": user.get("name", "") if user else "",
                "Referral ID": user.get("referralId", "") if user else "",
                "Type": txn.get("type", ""),
                "Amount": f"₹{txn.get('amount', 0)}",
                "Description": txn.get("description", "")
            })
        
        if format == "excel":
            headers = ["Date", "User", "Referral ID", "Type", "Amount", "Description"]
            output = generate_excel_report(report_data, headers, "Earnings Summary Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=earnings_report_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Date", "User", "Referral ID", "Type", "Amount"]
            pdf_data = [{k: v for k, v in item.items() if k != "Description"} for item in report_data]
            output = generate_pdf_report(pdf_data, headers, "Earnings Summary Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=earnings_report_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            total_earnings = sum([txn.get("amount", 0) for txn in transactions])
            return {
                "success": True,
                "data": report_data,
                "summary": {
                    "total": len(report_data),
                    "totalAmount": total_earnings
                }
            }
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/financial/income-breakdown")
async def get_income_breakdown_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get income breakdown by type"""
    try:
        start, end = parse_date_range(start_date, end_date)
        
        query = {"amount": {"$gt": 0}, "type": {"$in": ["REFERRAL_INCOME", "MATCHING_INCOME", "LEVEL_INCOME"]}}
        if start or end:
            query["createdAt"] = {}
            if start:
                query["createdAt"]["$gte"] = start
            if end:
                query["createdAt"]["$lte"] = end
        
        transactions = list(transactions_collection.find(query))
        
        # Group by type
        breakdown = {}
        for txn in transactions:
            income_type = txn.get("type", "UNKNOWN")
            if income_type not in breakdown:
                breakdown[income_type] = {"count": 0, "total": 0, "transactions": []}
            breakdown[income_type]["count"] += 1
            breakdown[income_type]["total"] += txn.get("amount", 0)
            breakdown[income_type]["transactions"].append(txn)
        
        report_data = []
        for income_type, data in breakdown.items():
            report_data.append({
                "Income Type": income_type.replace("_", " ").title(),
                "Transaction Count": data["count"],
                "Total Amount": f"₹{data['total']}"
            })
        
        if format == "excel":
            headers = ["Income Type", "Transaction Count", "Total Amount"]
            output = generate_excel_report(report_data, headers, "Income Breakdown Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=income_breakdown_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Income Type", "Transaction Count", "Total Amount"]
            output = generate_pdf_report(report_data, headers, "Income Breakdown Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=income_breakdown_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {"success": True, "data": report_data, "breakdown": breakdown}
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/financial/withdrawals")
async def get_withdrawals_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get withdrawals/payout history report"""
    try:
        start, end = parse_date_range(start_date, end_date)
        
        query = {}
        if start or end:
            query["createdAt"] = {}
            if start:
                query["createdAt"]["$gte"] = start
            if end:
                query["createdAt"]["$lte"] = end
        
        if status and status != "all":
            query["status"] = status.upper()
        
        withdrawals = list(withdrawals_collection.find(query))
        
        report_data = []
        for withdrawal in withdrawals:
            user = users_collection.find_one({"_id": ObjectId(withdrawal.get("userId"))})
            report_data.append({
                "Date": withdrawal.get("createdAt", datetime.now()).strftime("%d-%m-%Y") if withdrawal.get("createdAt") else "",
                "User": user.get("name", "") if user else "",
                "Referral ID": user.get("referralId", "") if user else "",
                "Amount": f"₹{withdrawal.get('amount', 0)}",
                "Status": withdrawal.get("status", ""),
                "Approved Date": withdrawal.get("approvedAt", datetime.now()).strftime("%d-%m-%Y") if withdrawal.get("approvedAt") else "N/A"
            })
        
        if format == "excel":
            headers = ["Date", "User", "Referral ID", "Amount", "Status", "Approved Date"]
            output = generate_excel_report(report_data, headers, "Withdrawals Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=withdrawals_report_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Date", "User", "Referral ID", "Amount", "Status"]
            pdf_data = [{k: v for k, v in item.items() if k != "Approved Date"} for item in report_data]
            output = generate_pdf_report(pdf_data, headers, "Withdrawals Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=withdrawals_report_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            total_amount = sum([w.get("amount", 0) for w in withdrawals])
            return {
                "success": True,
                "data": report_data,
                "summary": {
                    "total": len(report_data),
                    "totalAmount": total_amount
                }
            }
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/financial/topups")
async def get_topups_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get topups history report"""
    try:
        start, end = parse_date_range(start_date, end_date)
        
        query = {}
        if start or end:
            query["createdAt"] = {}
            if start:
                query["createdAt"]["$gte"] = start
            if end:
                query["createdAt"]["$lte"] = end
        
        topups = list(topups_collection.find(query))
        
        report_data = []
        for topup in topups:
            user = users_collection.find_one({"_id": ObjectId(topup.get("userId"))})
            report_data.append({
                "Date": topup.get("createdAt", datetime.now()).strftime("%d-%m-%Y") if topup.get("createdAt") else "",
                "User": user.get("name", "") if user else "",
                "Referral ID": user.get("referralId", "") if user else "",
                "Amount": f"₹{topup.get('amount', 0)}",
                "Status": topup.get("status", ""),
                "Payment Method": topup.get("paymentMethod", "")
            })
        
        if format == "excel":
            headers = ["Date", "User", "Referral ID", "Amount", "Status", "Payment Method"]
            output = generate_excel_report(report_data, headers, "Topups Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=topups_report_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Date", "User", "Referral ID", "Amount", "Status", "Payment Method"]
            output = generate_pdf_report(report_data, headers, "Topups Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=topups_report_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            total_amount = sum([t.get("amount", 0) for t in topups])
            return {
                "success": True,
                "data": report_data,
                "summary": {
                    "total": len(report_data),
                    "totalAmount": total_amount
                }
            }
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/financial/business")
async def get_business_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get daily/weekly/monthly business report"""
    try:
        start, end = parse_date_range(start_date, end_date)
        
        if not start:
            start = datetime.now() - timedelta(days=30)
        if not end:
            end = datetime.now()
        
        # Generate daily reports
        daily_reports = []
        current_date = start
        while current_date <= end:
            day_end = current_date.replace(hour=23, minute=59, second=59)
            
            # New registrations
            new_users = users_collection.count_documents({
                "role": "user",
                "createdAt": {"$gte": current_date, "$lte": day_end}
            })
            
            # Topups
            topups_pipeline = [
                {"$match": {"status": "APPROVED", "approvedAt": {"$gte": current_date, "$lte": day_end}}},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ]
            topups_result = list(topups_collection.aggregate(topups_pipeline))
            topups_amount = topups_result[0]["total"] if topups_result else 0
            
            # Payouts
            payouts_pipeline = [
                {"$match": {"status": "APPROVED", "approvedAt": {"$gte": current_date, "$lte": day_end}}},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ]
            payouts_result = list(withdrawals_collection.aggregate(payouts_pipeline))
            payouts_amount = payouts_result[0]["total"] if payouts_result else 0
            
            daily_reports.append({
                "Date": current_date.strftime("%d-%m-%Y"),
                "New Users": new_users,
                "Topups": f"₹{topups_amount}",
                "Payouts": f"₹{payouts_amount}",
                "Net Business": f"₹{topups_amount - payouts_amount}"
            })
            
            current_date += timedelta(days=1)
        
        if format == "excel":
            headers = ["Date", "New Users", "Topups", "Payouts", "Net Business"]
            output = generate_excel_report(daily_reports, headers, "Daily Business Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=business_report_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Date", "New Users", "Topups", "Payouts", "Net Business"]
            output = generate_pdf_report(daily_reports, headers, "Daily Business Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=business_report_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {"success": True, "data": daily_reports, "total": len(daily_reports)}
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# TEAM/NETWORK REPORTS

@app.get("/api/admin/reports/team/structure")
async def get_team_structure_report(
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get complete team structure report"""
    try:
        teams = list(teams_collection.find({}))
        
        report_data = []
        for team in teams:
            user = users_collection.find_one({"referralId": team.get("userId")})
            sponsor = users_collection.find_one({"referralId": team.get("sponsorId")})
            
            if user:
                report_data.append({
                    "User ID": team.get("userId", ""),
                    "User Name": user.get("name", ""),
                    "Sponsor ID": team.get("sponsorId", ""),
                    "Sponsor Name": sponsor.get("name", "") if sponsor else "",
                    "Placement": team.get("placement", ""),
                    "Joined Date": user.get("createdAt", datetime.now()).strftime("%d-%m-%Y") if user.get("createdAt") else ""
                })
        
        if format == "excel":
            headers = ["User ID", "User Name", "Sponsor ID", "Sponsor Name", "Placement", "Joined Date"]
            output = generate_excel_report(report_data, headers, "Team Structure Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=team_structure_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["User ID", "User Name", "Sponsor ID", "Sponsor Name", "Placement"]
            pdf_data = [{k: v for k, v in item.items() if k != "Joined Date"} for item in report_data]
            output = generate_pdf_report(pdf_data, headers, "Team Structure Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=team_structure_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {"success": True, "data": report_data, "total": len(report_data)}
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/team/downline")
async def get_downline_report(
    referral_id: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get downline summary for a specific user or all users"""
    try:
        if referral_id:
            users_to_check = [users_collection.find_one({"referralId": referral_id})]
        else:
            users_to_check = list(users_collection.find({"role": "user"}))
        
        report_data = []
        for user in users_to_check:
            if not user:
                continue
                
            user_id = user.get("referralId", "")
            
            # Count direct downline
            direct_count = teams_collection.count_documents({"sponsorId": user_id})
            
            # Get all downline recursively
            def get_all_downline(sponsor_id, visited=None):
                if visited is None:
                    visited = set()
                if sponsor_id in visited:
                    return []
                visited.add(sponsor_id)
                
                direct = list(teams_collection.find({"sponsorId": sponsor_id}))
                all_downline = direct.copy()
                for member in direct:
                    all_downline.extend(get_all_downline(member.get("userId"), visited))
                return all_downline
            
            total_downline = len(get_all_downline(user_id))
            
            report_data.append({
                "Referral ID": user_id,
                "Name": user.get("name", ""),
                "Direct Downline": direct_count,
                "Total Downline": total_downline,
                "Status": "Active" if user.get("isActive", False) else "Inactive"
            })
        
        if format == "excel":
            headers = ["Referral ID", "Name", "Direct Downline", "Total Downline", "Status"]
            output = generate_excel_report(report_data, headers, "Downline Summary Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=downline_summary_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Referral ID", "Name", "Direct Downline", "Total Downline", "Status"]
            output = generate_pdf_report(report_data, headers, "Downline Summary Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=downline_summary_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {"success": True, "data": report_data, "total": len(report_data)}
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/team/binary-tree")
async def get_binary_tree_export(
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Export binary tree data"""
    try:
        teams = list(teams_collection.find({}))
        
        report_data = []
        for team in teams:
            user = users_collection.find_one({"referralId": team.get("userId")})
            if user:
                report_data.append({
                    "User ID": team.get("userId", ""),
                    "User Name": user.get("name", ""),
                    "Sponsor ID": team.get("sponsorId", ""),
                    "Position": team.get("placement", ""),
                    "Left Side Count": team.get("leftCount", 0),
                    "Right Side Count": team.get("rightCount", 0),
                    "Status": "Active" if user.get("isActive", False) else "Inactive"
                })
        
        if format == "excel":
            headers = ["User ID", "User Name", "Sponsor ID", "Position", "Left Side Count", "Right Side Count", "Status"]
            output = generate_excel_report(report_data, headers, "Binary Tree Data Export")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=binary_tree_data_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["User ID", "User Name", "Sponsor ID", "Position", "Left Count", "Right Count"]
            pdf_data = []
            for item in report_data:
                pdf_data.append({
                    "User ID": item["User ID"],
                    "User Name": item["User Name"],
                    "Sponsor ID": item["Sponsor ID"],
                    "Position": item["Position"],
                    "Left Count": item["Left Side Count"],
                    "Right Count": item["Right Side Count"]
                })
            output = generate_pdf_report(pdf_data, headers, "Binary Tree Data Export")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=binary_tree_data_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {"success": True, "data": report_data, "total": len(report_data)}
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ANALYTICS REPORTS

@app.get("/api/admin/reports/analytics/registrations")
async def get_registrations_trend(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get daily registrations trend"""
    try:
        start, end = parse_date_range(start_date, end_date)
        
        if not start:
            start = datetime.now() - timedelta(days=30)
        if not end:
            end = datetime.now()
        
        report_data = []
        current_date = start
        while current_date <= end:
            day_end = current_date.replace(hour=23, minute=59, second=59)
            
            count = users_collection.count_documents({
                "role": "user",
                "createdAt": {"$gte": current_date, "$lte": day_end}
            })
            
            report_data.append({
                "Date": current_date.strftime("%d-%m-%Y"),
                "New Registrations": count
            })
            
            current_date += timedelta(days=1)
        
        if format == "excel":
            headers = ["Date", "New Registrations"]
            output = generate_excel_report(report_data, headers, "Daily Registrations Trend")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=registrations_trend_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Date", "New Registrations"]
            output = generate_pdf_report(report_data, headers, "Daily Registrations Trend")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=registrations_trend_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            total_registrations = sum([r["New Registrations"] for r in report_data])
            return {
                "success": True,
                "data": report_data,
                "summary": {"total": total_registrations}
            }
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/analytics/plan-distribution")
async def get_plan_distribution_report(
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get plan distribution analysis"""
    try:
        plans = list(plans_collection.find({}))
        
        report_data = []
        total_users_with_plan = 0
        
        for plan in plans:
            plan_id_str = str(plan["_id"])
            count = users_collection.count_documents({
                "$or": [
                    {"currentPlan": plan_id_str},
                    {"currentPlan": plan["_id"]},
                    {"currentPlan": plan["name"]}
                ]
            })
            total_users_with_plan += count
            
            report_data.append({
                "Plan Name": plan.get("name", ""),
                "Price": f"₹{plan.get('price', 0)}",
                "User Count": count,
                "Revenue": f"₹{plan.get('price', 0) * count}"
            })
        
        # Add no plan users
        no_plan_count = users_collection.count_documents({
            "role": "user",
            "$or": [
                {"currentPlan": None},
                {"currentPlan": ""},
                {"currentPlan": {"$exists": False}}
            ]
        })
        
        report_data.append({
            "Plan Name": "No Plan",
            "Price": "₹0",
            "User Count": no_plan_count,
            "Revenue": "₹0"
        })
        
        if format == "excel":
            headers = ["Plan Name", "Price", "User Count", "Revenue"]
            output = generate_excel_report(report_data, headers, "Plan Distribution Analysis")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=plan_distribution_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Plan Name", "Price", "User Count", "Revenue"]
            output = generate_pdf_report(report_data, headers, "Plan Distribution Analysis")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=plan_distribution_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {"success": True, "data": report_data, "total": len(report_data)}
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/reports/analytics/growth")
async def get_growth_statistics(
    format: str = "json",
    current_admin: dict = Depends(get_current_admin)
):
    """Get growth statistics"""
    try:
        # Monthly growth for last 12 months
        report_data = []
        
        for i in range(11, -1, -1):
            month_start = (datetime.now().replace(day=1) - timedelta(days=i*30)).replace(day=1, hour=0, minute=0, second=0)
            if i == 0:
                month_end = datetime.now()
            else:
                month_end = (datetime.now().replace(day=1) - timedelta(days=(i-1)*30)).replace(day=1, hour=0, minute=0, second=0)
            
            new_users = users_collection.count_documents({
                "role": "user",
                "createdAt": {"$gte": month_start, "$lt": month_end}
            })
            
            total_users = users_collection.count_documents({
                "role": "user",
                "createdAt": {"$lt": month_end}
            })
            
            # Calculate revenue for the month
            topups_pipeline = [
                {"$match": {"status": "APPROVED", "approvedAt": {"$gte": month_start, "$lt": month_end}}},
                {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
            ]
            topups_result = list(topups_collection.aggregate(topups_pipeline))
            revenue = topups_result[0]["total"] if topups_result else 0
            
            report_data.append({
                "Month": month_start.strftime("%B %Y"),
                "New Users": new_users,
                "Total Users": total_users,
                "Revenue": f"₹{revenue}"
            })
        
        if format == "excel":
            headers = ["Month", "New Users", "Total Users", "Revenue"]
            output = generate_excel_report(report_data, headers, "Growth Statistics Report")
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=growth_statistics_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.xlsx"}
            )
        elif format == "pdf":
            headers = ["Month", "New Users", "Total Users", "Revenue"]
            output = generate_pdf_report(report_data, headers, "Growth Statistics Report")
            return StreamingResponse(
                output,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=growth_statistics_{datetime.now(IST).strftime('%Y%m%d_%H%M%S')}.pdf"}
            )
        else:
            return {"success": True, "data": report_data, "total": len(report_data)}
    
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/admin/calculate-daily-matching")
async def calculate_daily_matching_income(current_admin: dict = Depends(get_current_admin)):
    """
    Calculate matching income for all users at end of day
    This should be called once per day (manually or via cron job)
    """
    try:
        # Get all users with active plans (including admin)
        users = list(users_collection.find({
            "isActive": True,
            "currentPlan": {"$ne": None}
        }))
        
        total_processed = 0
        total_income_paid = 0
        results = []
        
        for user in users:
            user_id = str(user["_id"])
            left_pv = user.get("leftPV", 0)
            right_pv = user.get("rightPV", 0)
            
            # Skip if no PV on either side
            if left_pv == 0 or right_pv == 0:
                continue
            
            # Calculate matching income
            try:
                # Get plan details
                plan = None
                plan_id = user.get("currentPlanId")
                plan_value = user.get("currentPlan")
                
                if plan_id:
                    # Try by currentPlanId
                    plan = plans_collection.find_one({"_id": ObjectId(plan_id)})
                elif plan_value:
                    # Try by name first
                    plan = plans_collection.find_one({"name": plan_value})
                    if not plan:
                        # Try as ObjectId string
                        try:
                            plan = plans_collection.find_one({"_id": ObjectId(plan_value)})
                        except:
                            pass
                
                if not plan:
                    continue
                
                daily_capping = plan.get("dailyCapping", 500)
                matching_income_rate = 25  # ₹25 per PV
                
                # Calculate matching PV
                matched_pv = min(left_pv, right_pv)
                
                # Check daily capping
                today_date = datetime.now(IST).replace(hour=0, minute=0, second=0, microsecond=0)
                last_matching_date = user.get("lastMatchingDate")
                
                # Reset daily PV if new day
                if not last_matching_date or last_matching_date.replace(hour=0, minute=0, second=0, microsecond=0) != today_date:
                    daily_pv_used = 0
                else:
                    daily_pv_used = user.get("dailyPVUsed", 0)
                
                # Calculate maximum PV allowed today
                max_pv_per_day = daily_capping // matching_income_rate
                remaining_pv_today = max_pv_per_day - daily_pv_used
                
                if remaining_pv_today <= 0:
                    continue  # Daily limit reached
                
                # Today's PV = min(matched_pv, remaining_pv_today)
                today_pv = min(matched_pv, remaining_pv_today)
                
                if today_pv <= 0:
                    continue
                
                # Calculate income
                income = today_pv * matching_income_rate
                
                # Update wallet
                wallets_collection.update_one(
                    {"userId": user_id},
                    {
                        "$inc": {
                            "balance": income,
                            "totalEarnings": income
                        },
                        "$set": {"updatedAt": datetime.now(IST)}
                    }
                )
                
                # Create transaction
                transactions_collection.insert_one({
                    "userId": user_id,
                    "type": "MATCHING_INCOME",
                    "amount": income,
                    "description": f"Daily binary matching income - {today_pv} PV @ ₹{matching_income_rate}/PV",
                    "pv": today_pv,
                    "status": "COMPLETED",
                    "createdAt": datetime.now(IST)
                })
                
                # Flush matched PV from both sides
                users_collection.update_one(
                    {"_id": user["_id"]},
                    {
                        "$inc": {
                            "leftPV": -today_pv,
                            "rightPV": -today_pv,
                            "totalPV": today_pv
                        },
                        "$set": {
                            "lastMatchingDate": today_date,
                            "dailyPVUsed": daily_pv_used + today_pv,
                            "updatedAt": datetime.now(IST)
                        }
                    }
                )
                
                total_processed += 1
                total_income_paid += income
                
                results.append({
                    "userId": user_id,
                    "name": user.get("name"),
                    "referralId": user.get("referralId"),
                    "matchedPV": today_pv,
                    "income": income,
                    "leftPV_before": left_pv,
                    "rightPV_before": right_pv,
                    "leftPV_after": left_pv - today_pv,
                    "rightPV_after": right_pv - today_pv
                })
                
            except Exception as e:
                print(f"Error processing user {user_id}: {str(e)}")
                continue
        
        return {
            "success": True,
            "message": "Daily matching income calculated successfully",
            "summary": {
                "totalUsersProcessed": total_processed,
                "totalIncomePaid": total_income_paid,
                "date": datetime.now(IST).strftime("%Y-%m-%d")
            },
            "details": results
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    try:
        # Check MongoDB connection
        client.admin.command('ping')
        return {
            "status": "healthy",
            "database": "connected",
            "timestamp": get_ist_now().isoformat()
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "database": "disconnected",
            "error": str(e),
            "timestamp": get_ist_now().isoformat()
        }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="0.0.0.0", port=8001, reload=True)
